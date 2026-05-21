from __future__ import annotations

import csv
import ctypes
import hashlib
import json
import mimetypes
import queue
import shutil
import tempfile
import threading
import time
import uuid
import ast
import os
import re
import subprocess
import sys
import webbrowser
from dataclasses import dataclass
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, unquote, urlparse

import spectral_physics_api as spectral_api

APP_DIR = Path(__file__).resolve().parent
ROOT = APP_DIR.parent

PNG_DIR_NAMES = ("03_transmission_abs2_png", "03_transmission_png_abs2")
XLSX_DIR_NAMES = ("02_transmission_excel",)
FSP_DIR_NAMES = ("01_fsp", "01_supercell_fsp")
PARAM_SKIP_KEYS = {
    "status",
    "fsp",
    "fsp_file",
    "xlsx",
    "excel_file",
    "png",
    "png_file",
    "elapsed_s",
    "max_abs2",
    "max_wavelength_nm",
    "min_abs2",
    "min_wavelength_nm",
}
FDTD_RUNTIME_KEYS = ("SIMULATION_TIME_FS", "AUTO_SHUTOFF_MIN", "MESH_ACCURACY", "DT_STABILITY_FACTOR")
LEGACY_FDTD_RUNTIME_KEYS = ("SIMULATION_TIME_S",)


@dataclass(frozen=True)
class FileRef:
    id: str
    path: Path


FILES: dict[str, FileRef] = {}
JOBS: dict[str, dict] = {}
JOB_PROCS: dict[str, subprocess.Popen] = {}
JOBS_LOCK = threading.Lock()
STATE_DIR = APP_DIR / "runtime_state"
GENERATED_DIR = APP_DIR / "generated"
LOG_DIR = APP_DIR / "logs"
TEST_ARTIFACT_DIR = APP_DIR / "test_artifacts"
for _runtime_dir in (STATE_DIR, GENERATED_DIR, LOG_DIR, TEST_ARTIFACT_DIR):
    try:
        _runtime_dir.mkdir(exist_ok=True)
    except Exception:
        pass
SCAN_CACHE_FILE = STATE_DIR / "scan_cache.json"
LEGACY_SCAN_CACHE_FILE = APP_DIR / "scan_cache.json"
SPECTRAL_CONFIG_FILE = STATE_DIR / "spectral_config.json"
LEGACY_SPECTRAL_CONFIG_FILE = APP_DIR / "spectral_config.json"
VIEW_STATE_FILE = STATE_DIR / "view_state.json"
LEGACY_VIEW_STATE_FILE = APP_DIR / "view_state.json"
SCAN_CACHE_VERSION = 4
SCAN_CACHE_MAX_BYTES = 50 * 1024 * 1024
SCAN_CACHE_TTL_SECONDS = 7 * 24 * 3600
UNCONVERGED_FOLDER_NAME = "\u4e0d\u6536\u655b\u7ed3\u679c"
ARCHIVE_DIR_NAME = "\u65e7\u6587\u4ef6"
UNCONVERGED_QUALITY_NAME = "\u4e0d\u6536\u655b"
KEEP_LATEST_ACTIVE_RUNS = 1
SCAN_LOCK = threading.Lock()
SCAN_REFRESHING = False
SCAN_REFRESH_STARTED = 0.0
SCAN_REFRESH_FINISHED = 0.0



def is_under_root(path: Path) -> bool:
    try:
        path.resolve().relative_to(ROOT.resolve())
        return True
    except ValueError:
        return False


def to_rel(path: Path) -> str:
    try:
        return str(path.relative_to(ROOT))
    except ValueError:
        return str(path)


def safe_id(path: Path) -> str:
    rel = to_rel(path).replace("\\", "/").lower()
    return hashlib.sha1(rel.encode("utf-8", errors="replace")).hexdigest()[:20]


def read_text_guess(path: Path, limit: int | None = None) -> str:
    data = path.read_bytes()
    if limit:
        data = data[:limit]
    for enc in ("utf-8-sig", "utf-8", "gb18030", "cp936"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    text = read_text_guess(path)
    return list(csv.DictReader(text.splitlines()))


def register(path: Path) -> str:
    fid = safe_id(path)
    FILES[fid] = FileRef(fid, path)
    return fid


def files_snapshot() -> dict:
    return {fid: to_rel(ref.path) for fid, ref in FILES.items()}


def restore_files_snapshot(files: dict):
    FILES.clear()
    if not isinstance(files, dict):
        return
    for fid, rel in files.items():
        try:
            path = (ROOT / str(rel)).resolve()
            if is_under_root(path):
                FILES[str(fid)] = FileRef(str(fid), path)
        except Exception:
            continue


def load_scan_cache() -> dict | None:
    cache_file = SCAN_CACHE_FILE if SCAN_CACHE_FILE.exists() else LEGACY_SCAN_CACHE_FILE
    if not cache_file.exists():
        return None
    try:
        stat = cache_file.stat()
        if stat.st_size > SCAN_CACHE_MAX_BYTES:
            return None
        data = json.loads(cache_file.read_text(encoding="utf-8"))
        if data.get("version") != SCAN_CACHE_VERSION:
            return None
        saved_at = float(data.get("saved_at", 0) or 0)
        if saved_at and time.time() - saved_at > SCAN_CACHE_TTL_SECONDS:
            return None
        payload = data.get("payload")
        if not isinstance(payload, dict):
            return None
        restore_files_snapshot(data.get("files", {}))
        payload["cached"] = True
        payload["cache_time"] = data.get("saved_at", 0)
        return payload
    except Exception:
        return None


def save_scan_cache(payload: dict):
    try:
        data = {"version": SCAN_CACHE_VERSION, "saved_at": time.time(), "root": str(ROOT), "payload": payload, "files": files_snapshot()}
        tmp = SCAN_CACHE_FILE.with_suffix(".tmp")
        tmp.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
        tmp.replace(SCAN_CACHE_FILE)
    except Exception:
        pass


def recycle_folder(path: Path):
    if not path.is_dir() or not is_under_root(path) or not path.name.startswith("run_"):
        raise RuntimeError("Refusing to delete non-run folder")
    recycle_path(path)


def read_json_body(handler) -> dict:
    length = int(handler.headers.get("Content-Length", "0") or 0)
    if length <= 0:
        return {}
    raw = handler.rfile.read(length).decode("utf-8", errors="replace")
    try:
        data = json.loads(raw) if raw.strip() else {}
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"invalid JSON body: {exc}")
    return data if isinstance(data, dict) else {}


def normalize_override_payload(raw_overrides: dict) -> dict:
    if not isinstance(raw_overrides, dict):
        return {}
    normalized: dict[str, dict] = {}
    for key, value in raw_overrides.items():
        if not isinstance(value, dict):
            continue
        clean = {str(k): v for k, v in value.items() if v not in (None, "")}
        if not clean:
            continue
        key_text = str(key)
        if key_text == "*":
            normalized[key_text] = clean
            continue
        aliases = {key_text}
        try:
            aliases.add(str(Path(key_text).resolve()))
        except Exception:
            pass
        for alias in aliases:
            normalized[alias] = dict(clean)
    wildcard = normalized.get("*")
    if wildcard:
        for key, value in list(normalized.items()):
            if key == "*":
                continue
            merged = dict(wildcard)
            merged.update(value)
            normalized[key] = merged
    return normalized


def load_view_state() -> dict:
    view_file = VIEW_STATE_FILE if VIEW_STATE_FILE.exists() else LEGACY_VIEW_STATE_FILE
    if not view_file.exists():
        return {"viewed_runs": {}, "run_tags": {}}
    try:
        data = json.loads(view_file.read_text(encoding="utf-8"))
        if not isinstance(data, dict):
            return {"viewed_runs": {}, "run_tags": {}}
        data.setdefault("viewed_runs", {})
        data.setdefault("run_tags", {})
        return data
    except Exception:
        return {"viewed_runs": {}, "run_tags": {}}


def save_view_state(data: dict):
    VIEW_STATE_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def mark_run_viewed(path: Path):
    data = load_view_state()
    data.setdefault("viewed_runs", {})[to_rel(path)] = time.time()
    save_view_state(data)


def is_run_viewed(path: Path) -> bool:
    return to_rel(path) in load_view_state().get("viewed_runs", {})


def run_viewed_at(path: Path) -> float:
    try:
        return float(load_view_state().get("viewed_runs", {}).get(to_rel(path), 0) or 0)
    except Exception:
        return 0.0


def run_tags(path: Path) -> list[str]:
    tags = load_view_state().get("run_tags", {}).get(to_rel(path), [])
    if not isinstance(tags, list):
        return []
    return [str(tag) for tag in tags if str(tag).strip()]


def set_run_tags(path: Path, tags: list[str]) -> list[str]:
    clean: list[str] = []
    seen: set[str] = set()
    for tag in tags:
        text = str(tag).strip()
        if not text or text in seen:
            continue
        seen.add(text)
        clean.append(text)
    data = load_view_state()
    data.setdefault("run_tags", {})[to_rel(path)] = clean
    save_view_state(data)
    return clean


def recycle_path(path: Path):
    if os.name != "nt":
        raise RuntimeError("Recycle bin delete is only supported on Windows")
    if not path.exists() or not is_under_root(path):
        raise RuntimeError("Refusing to delete path outside root")

    class SHFILEOPSTRUCTW(ctypes.Structure):
        _fields_ = [
            ("hwnd", ctypes.c_void_p),
            ("wFunc", ctypes.c_uint),
            ("pFrom", ctypes.c_wchar_p),
            ("pTo", ctypes.c_wchar_p),
            ("fFlags", ctypes.c_ushort),
            ("fAnyOperationsAborted", ctypes.c_bool),
            ("hNameMappings", ctypes.c_void_p),
            ("lpszProgressTitle", ctypes.c_wchar_p),
        ]

    FO_DELETE = 0x0003
    FOF_ALLOWUNDO = 0x0040
    FOF_NOCONFIRMATION = 0x0010
    FOF_SILENT = 0x0004
    op = SHFILEOPSTRUCTW()
    op.wFunc = FO_DELETE
    op.pFrom = str(path.resolve()) + "\0\0"
    op.fFlags = FOF_ALLOWUNDO | FOF_NOCONFIRMATION | FOF_SILENT
    result = ctypes.windll.shell32.SHFileOperationW(ctypes.byref(op))
    if result != 0 or op.fAnyOperationsAborted:
        raise RuntimeError(f"Windows recycle operation failed: {result}")


def unique_destination(path: Path) -> Path:
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    for i in range(1, 1000):
        candidate = path.with_name(f"{stem}_{i:03d}{suffix}")
        if not candidate.exists():
            return candidate
    raise RuntimeError(f"cannot find unique destination for {path}")


def run_mode_from_name(path: Path) -> str:
    name = path.name.lower()
    if name.startswith("run_full") or "full" in name:
        return "full"
    if name.startswith("run_test") or "test" in name:
        return "test"
    if name.startswith("run_preview") or "preview" in name:
        return "preview"
    return "unknown"


def is_archived_or_collected(path: Path) -> bool:
    return ARCHIVE_DIR_NAME in path.parts or UNCONVERGED_FOLDER_NAME in path.parts


def archive_quality_destination(run_dir: Path, quality: str) -> Path:
    mode = run_mode_from_name(run_dir)
    parent = run_dir.parent / ARCHIVE_DIR_NAME / mode / quality
    parent.mkdir(parents=True, exist_ok=True)
    return unique_destination(parent / run_dir.name)


def run_refs_from_ids(ids: list[str]) -> list[Path]:
    paths: list[Path] = []
    seen: set[Path] = set()
    for fid in ids:
        ref = FILES.get(str(fid))
        if not ref or not ref.path.exists() or not ref.path.is_dir():
            continue
        path = ref.path.resolve()
        if path in seen:
            continue
        if not is_under_root(path):
            continue
        seen.add(path)
        paths.append(path)
    return paths


def move_unconverged_runs(run_ids: list[str] | None = None) -> dict:
    moved = []
    skipped = []
    if run_ids:
        run_dirs = run_refs_from_ids(run_ids)
    else:
        # Fallback: scan current runs from the live cache
        data = scan_runs_uncached()
        run_dirs = []
        for run in data.get("runs", []):
            if not any(item.get("unconverged") for item in run.get("items", [])):
                continue
            ref = FILES.get(run.get("id", ""))
            if ref and ref.path.is_dir():
                run_dirs.append(ref.path)

    if not run_dirs:
        return {"ok": True, "moved": [], "skipped": [], "info": "no unconverged run needs moving"}

    for run_dir in run_dirs:
        try:
            if is_archived_or_collected(run_dir):
                skipped.append({"path": to_rel(run_dir), "reason": "already archived"})
                continue
            dest = archive_quality_destination(run_dir, UNCONVERGED_QUALITY_NAME)
            if not is_under_root(run_dir) or not is_under_root(dest):
                skipped.append({"path": to_rel(run_dir), "reason": "outside root"})
                continue
            shutil.move(str(run_dir), str(dest))
            moved.append({"from": to_rel(run_dir), "to": to_rel(dest)})
        except PermissionError as exc:
            skipped.append({"path": to_rel(run_dir), "reason": f"file in use: {exc}"})
        except OSError as exc:
            skipped.append({"path": to_rel(run_dir), "reason": str(exc)})
    return {"ok": True, "moved": moved, "skipped": skipped}


def parse_literal_assignments(script_path: Path) -> tuple[dict, str]:
    text = script_path.read_text(encoding="utf-8", errors="replace")
    try:
        tree = ast.parse(text)
    except SyntaxError:
        return {}, text
    values = {}
    for node in tree.body:
        if isinstance(node, ast.Assign) and len(node.targets) == 1 and isinstance(node.targets[0], ast.Name):
            name = node.targets[0].id
            if name == "CONFIG":
                break
            if not name.isupper():
                continue
            try:
                values[name] = ast.literal_eval(node.value)
            except Exception:
                pass
    return values, text


def script_param_groups(script_path: Path) -> dict:
    values, _ = parse_literal_assignments(script_path)
    scan = []

    def add(prefix, unit, s, e, st):
        if s in values and e in values and st in values:
            scan.append({"prefix": prefix, "unit": unit, "start_key": s, "end_key": e, "step_key": st, "start": values[s], "end": values[e], "step": values[st]})

    for unit in ("NM", "M", "DEG"):
        add("SCAN", unit, f"START_{unit}", f"END_{unit}", f"STEP_{unit}")
    add("SCAN", "RAW", "START", "END", "STEP")
    for key in values:
        m = re.match(r"(.+)_START_(NM|M|DEG)$", key)
        if m:
            prefix, unit = m.group(1), m.group(2)
            add(prefix, unit, key, f"{prefix}_STOP_{unit}", f"{prefix}_STEP_{unit}")
    runtime = {key: values[key] for key in FDTD_RUNTIME_KEYS if key in values}
    if "SIMULATION_TIME_FS" not in values:
        runtime.update({key: values[key] for key in LEGACY_FDTD_RUNTIME_KEYS if key in values})
    return {"scan": scan, "runtime": runtime, "values": values}


def replace_assignments(text: str, replacements: dict) -> str:
    if "SIMULATION_TIME_FS" in replacements and "SIMULATION_TIME_S" not in replacements:
        try:
            replacements = dict(replacements)
            replacements["SIMULATION_TIME_S"] = float(replacements["SIMULATION_TIME_FS"]) * 1e-15
        except (TypeError, ValueError):
            pass
    config_key_aliases = {
        "SIMULATION_TIME_FS": ("simulation_time_fs", "SIMULATION_TIME_FS"),
        "SIMULATION_TIME_S": ("simulation_time_s", "SIMULATION_TIME_S"),
        "AUTO_SHUTOFF_MIN": ("auto_shutoff_min", "AUTO_SHUTOFF_MIN"),
        "MESH_ACCURACY": ("mesh_accuracy", "MESH_ACCURACY"),
        "DT_STABILITY_FACTOR": ("dt_stability_factor", "DT_STABILITY_FACTOR"),
    }
    pending_config_updates = {}
    for name, value in replacements.items():
        if value in (None, ""):
            continue
        try:
            literal = repr(float(value))
        except (TypeError, ValueError):
            literal = repr(value)
        pattern = re.compile(r"^({}\s*=\s*)(.+?)\s*$".format(re.escape(name)), re.M)
        text, _ = pattern.subn(r"\g<1>{}".format(literal), text, count=1)
        kw_pattern = re.compile(r"(\b{}\s*=\s*)([^,\n\)]+)".format(re.escape(name)))
        text, _ = kw_pattern.subn(r"\g<1>{}".format(literal), text, count=1)
        config_keys = config_key_aliases.get(name, ())
        for config_key in config_keys:
            dict_pattern = re.compile(r"((?:['\"]{}['\"])\s*:\s*)([^,\n\}}]+)".format(re.escape(config_key)))
            text, _ = dict_pattern.subn(r"\g<1>{}".format(literal), text, count=1)
            pending_config_updates[config_key] = value
    if pending_config_updates and "CONFIG" in text:
        payload = ", ".join("{!r}: {!r}".format(k, v) for k, v in pending_config_updates.items())
        injection = "\n# Runtime overrides injected by server.py\nCONFIG.update({%s})\n" % payload
        marker = 'if __name__ == "__main__":'
        if marker in text:
            text = text.replace(marker, injection + "\n" + marker, 1)
        else:
            text += injection
    return text

def temporary_script(script_path: Path, replacements: dict) -> Path:
    if not replacements:
        return script_path
    _, text = parse_literal_assignments(script_path)
    text = replace_assignments(text, replacements)
    temp = script_path.parent / ("_web_temp_{}_{}".format(int(time.time()), script_path.name))
    temp.write_text(text, encoding="utf-8")
    return temp


def useful_output_line(line: str) -> bool:
    low = line.lower()
    noisy = ("findfont", "openblas", "userwarning", "deprecated", "loaded more than 1 dll")
    if any(token in low for token in noisy):
        return False
    return bool(line.strip())



def command_for_script(script: Path, mode: str) -> list[str]:
    cmd = [sys.executable, str(script)]
    if mode == "ask":
        return cmd
    try:
        probe = script.read_text(encoding="utf-8", errors="replace")
    except Exception:
        probe = ""
    for module_name in re.findall(r"from\s+([A-Za-z_][A-Za-z0-9_]*_common)\s+import\s+run", probe):
        for candidate in (script.parent / (module_name + ".py"), script.parent.parent / (module_name + ".py"), script.parent.parent.parent / (module_name + ".py"), ROOT / (module_name + ".py")):
            if candidate.exists():
                probe += "\n" + candidate.read_text(encoding="utf-8", errors="replace")
                break
    if '"--mode"' in probe or "'--mode'" in probe:
        return cmd + ["--mode", mode]
    if '"--test-run"' in probe or "'--test-run'" in probe or '"--full-run"' in probe or "'--full-run'" in probe:
        return cmd + {"preview": ["--preview"], "test": ["--test-run"], "full": ["--full-run"]}.get(mode, [])
    if '"--test"' in probe or "'--test'" in probe or '"--full"' in probe or "'--full'" in probe:
        return cmd + {"preview": ["--preview"], "test": ["--test"], "full": ["--full"]}.get(mode, [])
    return cmd


def job_template(job_id: str, title: str, cmd: list[str] | None = None, cwd: Path | str | None = None) -> dict:
    return {
        "id": job_id,
        "title": title,
        "cmd": cmd or [],
        "cwd": str(cwd or ROOT),
        "lines": [],
        "returncode": None,
        "running": True,
        "started": time.time(),
        "pid": None,
        "total": 0,
        "current": 0,
        "progress": 0,
        "current_title": "",
        "current_detail": "",
        "last_output": "",
        "stop_requested": False,
    }


def update_job_from_line(job: dict, line: str):
    job["last_output"] = line

    m = re.search(r"输出此次结果目录[:：]\s*(.+)$", line)
    if m:
        remember_job_result_path(job, m.group(1).strip())

    for match in re.finditer(r"[A-Za-z]:\\[^\r\n]*?\\run_(?:full|test|preview)_[^\r\n]*", line, re.I):
        remember_job_result_path(job, match.group(0).strip())

    m = re.search(r"\u6700\u7ec8\u5c06\u8fd0\u884c\s+(\d+)\s+\u4e2a\u811a\u672c", line)
    if m:
        job["total"] = int(m.group(1))

    m = re.search(r"\u51c6\u5907(?:\u987a\u5e8f|\u5e76\u884c)\u542f\u52a8\s+(\d+)\s+\u4e2a\u811a\u672c", line)
    if m:
        job["total"] = int(m.group(1))

    m = re.search(r"\u542f\u52a8\s+\[(\d+)\]\s+(.+?)(?:\uff1a|:)", line)
    if m:
        current = int(job.get("current", 0) or 0) + 1
        job["current"] = current
        if not job.get("total"):
            job["total"] = current
        job["current_title"] = m.group(2).strip()
        job["current_detail"] = ""

    m = re.search(r"\[(\d+)/\s*(\d+)\]\s+(?:\u542f\u52a8|start):\s*(.+)", line, re.I)
    if m:
        job["current"] = int(m.group(1))
        job["total"] = int(m.group(2))
        job["current_title"] = m.group(3).strip()
        job["current_detail"] = ""

    m = re.search(r"\[(\d+)\s+OUT\]\s+(.+)", line)
    if m:
        detail = m.group(2).strip()
        if detail:
            job["current_detail"] = detail

    detail_keys = (
        "\u5f00\u59cb\u4eff\u771f",
        "\u5f53\u524d\u7528\u6237\u4fee\u6539\u533a\u5f71\u54cd",
        "\u626b\u63cf\u8ba1\u5212\u5df2\u4fdd\u5b58",
        "\u9884\u89c8\u6a21\u5f0f\u7ed3\u675f",
        "\u5355\u6b21\u4eff\u771f\u65f6\u95f4\u4e0a\u9650",
    )
    if any(key in line for key in detail_keys):
        job["current_detail"] = line

    total = int(job.get("total") or 0)
    current = int(job.get("current") or 0)
    if total > 0:
        job["progress"] = max(0, min(99 if job.get("running") else 100, round(current * 100 / total)))

    if "\u5168\u90e8\u4efb\u52a1\u7ed3\u675f" in line or "\u5168\u90e8\u5b8c\u6210" in line:
        job["progress"] = 100

    if "完成并保存" in line or "保存" in line and ("max" in line or "min" in line):
        refresh_job_latest_spectrum(job)


def latest_spectrum_from_run_dir(run_dir: Path) -> dict | None:
    if not run_dir.exists() or not run_dir.is_dir() or not is_under_root(run_dir):
        return None
    candidates: list[Path] = []
    for dirname in XLSX_DIR_NAMES:
        xlsx_dir = run_dir / dirname
        if xlsx_dir.is_dir():
            candidates.extend(p for p in xlsx_dir.glob("*.xlsx") if p.is_file())
    if not candidates:
        candidates.extend(p for p in run_dir.rglob("*.xlsx") if p.is_file() and is_under_root(p))
    if not candidates:
        return None
    latest = max(candidates, key=lambda p: p.stat().st_mtime)
    try:
        data = xlsx_series(latest)
    except Exception:
        return None
    points = data.get("points") or []
    if not points:
        return None
    max_points = 180
    if len(points) > max_points:
        step = max(1, len(points) // max_points)
        sampled = points[::step][:max_points]
    else:
        sampled = points
    return {
        "name": latest.name,
        "path": to_rel(latest),
        "points": sampled,
        "point_count": len(points),
        "metrics": data.get("metrics") or {},
        "columns": data.get("columns") or [],
    }


def estimate_total_from_command(cmd: list[str]) -> int:
    try:
        args = [str(x) for x in (cmd or [])]
        if "--ids" in args:
            raw = args[args.index("--ids") + 1]
            total = 0
            for part in raw.replace("\uff0c", ",").split(","):
                part = part.strip()
                if not part:
                    continue
                if "-" in part:
                    a, b = [int(x.strip()) for x in part.split("-", 1)]
                    total += abs(b - a) + 1
                else:
                    int(part)
                    total += 1
            return total
        if "--all" in args:
            return len(discover_all_scripts())
    except Exception:
        return 0
    return 0


def run_dir_from_any_path(path: Path) -> Path | None:
    try:
        path = path.resolve()
    except Exception:
        return None
    parts = list(path.parts)
    for idx in range(len(parts) - 1, -1, -1):
        if str(parts[idx]).lower().startswith("run_"):
            candidate = Path(*parts[: idx + 1])
            if candidate.exists() and candidate.is_dir() and is_under_root(candidate):
                return candidate
    if path.exists() and path.is_dir() and path.name.lower().startswith("run_") and is_under_root(path):
        return path
    return None


def remember_job_result_path(job: dict, raw_path: str):
    text = str(raw_path or "").strip().strip("\"'")
    if not text:
        return
    # Trim common trailing status fragments while preserving Windows paths with spaces.
    text = re.split(r"\s+\|\s+|\s+;\s+", text, maxsplit=1)[0].strip()
    try:
        run_dir = run_dir_from_any_path(Path(text))
    except Exception:
        run_dir = None
    if run_dir:
        job["current_result_dir"] = str(run_dir)


def latest_run_dir_for_job(job: dict) -> Path | None:
    raw = job.get("current_result_dir")
    if raw:
        try:
            run_dir = Path(raw)
            if run_dir.exists() and run_dir.is_dir() and is_under_root(run_dir):
                return run_dir
        except Exception:
            pass
    started = float(job.get("started") or 0)
    candidates: list[Path] = []
    for pattern in ("run_full_*", "run_test_*", "run_preview_*"):
        for path in ROOT.rglob(pattern):
            try:
                if not path.is_dir() or ARCHIVE_DIR_NAME in path.parts or UNCONVERGED_FOLDER_NAME in path.parts:
                    continue
                if started and path.stat().st_mtime + 30 < started:
                    continue
                if any((path / dirname).is_dir() for dirname in XLSX_DIR_NAMES) or any(path.rglob("*.xlsx")):
                    candidates.append(path)
            except Exception:
                continue
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def refresh_job_latest_spectrum(job: dict):
    now = time.time()
    if now - float(job.get("latest_spectrum_checked_at") or 0) < 2.5:
        return
    job["latest_spectrum_checked_at"] = now
    run_dir = latest_run_dir_for_job(job)
    if not run_dir:
        return
    job["current_result_dir"] = str(run_dir)
    try:
        spectrum = latest_spectrum_from_run_dir(run_dir)
    except Exception:
        spectrum = None
    if spectrum:
        job["latest_spectrum"] = spectrum

def kill_process_tree(pid: int):
    if not pid:
        return
    if os.name == "nt":
        subprocess.call(["taskkill", "/PID", str(pid), "/T", "/F"], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    else:
        try:
            os.kill(pid, 15)
        except Exception:
            pass


def start_background_job(cmd: list[str], cwd: Path, title: str, temp_paths: list[Path] | None = None) -> str:
    job_id = uuid.uuid4().hex[:12]
    job = job_template(job_id, title, cmd, cwd)
    job["total"] = estimate_total_from_command(cmd)
    with JOBS_LOCK:
        JOBS[job_id] = job

    def worker():
        proc = None
        try:
            env = os.environ.copy()
            env["PYTHONIOENCODING"] = "utf-8:replace"
            env["PYTHONUNBUFFERED"] = "1"
            proc = subprocess.Popen(cmd, cwd=str(cwd), stdout=subprocess.PIPE, stderr=subprocess.STDOUT, stdin=subprocess.DEVNULL, text=True, encoding="utf-8", errors="replace", bufsize=1, env=env)
            with JOBS_LOCK:
                job["pid"] = proc.pid
                JOB_PROCS[job_id] = proc
            for line in proc.stdout or []:
                line = line.rstrip("\r\n")
                if useful_output_line(line):
                    with JOBS_LOCK:
                        job["lines"].append(line)
                        job["lines"] = job["lines"][-20000:]
                        update_job_from_line(job, line)
            proc.wait()
            with JOBS_LOCK:
                job["returncode"] = proc.returncode
        except Exception as exc:
            with JOBS_LOCK:
                job["lines"].append("ERROR: {}".format(exc))
                job["returncode"] = -1
        finally:
            with JOBS_LOCK:
                JOB_PROCS.pop(job_id, None)
                job["running"] = False
                if job.get("returncode") == 0:
                    job["progress"] = 100
            for path in temp_paths or []:
                try:
                    if path.exists() and path.name.startswith("_web_temp_"):
                        path.unlink()
                except Exception:
                    pass

    thread = threading.Thread(target=worker, daemon=True)
    thread.start()
    return job_id


def start_operation_job(title: str, action) -> str:
    job_id = uuid.uuid4().hex[:12]
    job = job_template(job_id, title, [], ROOT)
    job["lines"] = ["operation job queued"]
    job["result"] = None
    with JOBS_LOCK:
        JOBS[job_id] = job

    def worker():
        try:
            result = action()
            with JOBS_LOCK:
                job["result"] = result
                moved = len(result.get("moved", []))
                deleted = len(result.get("deleted", []))
                skipped = len(result.get("skipped", []))
                job["current_title"] = title
                job["current_detail"] = f"moved {moved}, deleted {deleted}, skipped {skipped}"
                if moved:
                    job["lines"].append(f"\u5df2\u79fb\u52a8 {moved} \u4e2a run\u3002")
                if deleted:
                    job["lines"].append(f"\u5df2\u5220\u9664 {deleted} \u9879\u3002")
                if skipped:
                    job["lines"].append(f"\u8df3\u8fc7 {skipped} \u9879\uff0c\u591a\u6570\u662f\u6587\u4ef6\u5360\u7528\u6216\u5df2\u7ecf\u5904\u7406\u3002")
                for item in result.get("skipped", [])[:30]:
                    if isinstance(item, dict):
                        job["lines"].append(f"\u8df3\u8fc7\uff1a{item.get('path', '')} | {item.get('reason', '')}")
                    else:
                        job["lines"].append(f"\u8df3\u8fc7\uff1a{item}")
                job["returncode"] = 0
                job["progress"] = 100
        except Exception as exc:
            with JOBS_LOCK:
                job["lines"].append(f"ERROR: {exc}")
                job["returncode"] = -1
        finally:
            with JOBS_LOCK:
                job["running"] = False

    threading.Thread(target=worker, daemon=True).start()
    return job_id


def job_summary(job: dict) -> dict:
    return {
        "id": job.get("id", ""),
        "title": job.get("title", ""),
        "running": bool(job.get("running")),
        "returncode": job.get("returncode"),
        "started": job.get("started", 0),
        "pid": job.get("pid"),
        "total": job.get("total", 0),
        "current": job.get("current", 0),
        "progress": job.get("progress", 0),
        "current_title": job.get("current_title", ""),
        "current_detail": job.get("current_detail", ""),
        "last_output": job.get("last_output", ""),
        "line_count": len(job.get("lines", [])),
        "has_spectrum": bool(job.get("latest_spectrum")),
    }

def natural_key(value: str):
    return [int(part) if part.isdigit() else part.lower() for part in re.split(r"(\d+)", value)]


def run_sort_key(path: Path):
    return (path.stat().st_mtime if path.exists() else 0, natural_key(path.name))


def child_dir(parent: Path, names: tuple[str, ...]) -> Path | None:
    for name in names:
        path = parent / name
        if path.exists() and path.is_dir():
            return path
    return None


def first_note(run_dir: Path) -> Path | None:
    notes = sorted(run_dir.glob("*.md"), key=lambda p: natural_key(p.name))
    if not notes:
        return None
    for note in notes:
        if "\u8bf4\u660e" in note.name or "note" in note.name.lower():
            return note
    return notes[0]


def parse_reduction_path(note_text: str) -> str:
    if not note_text:
        return ""
    for line in note_text.splitlines():
        if "\u964d\u7fa4" not in line:
            continue
        cleaned = line.strip().lstrip("-*# ").strip()
        match = re.search(r"\u964d\u7fa4(?:\u8def\u5f84|\u8def\u7ebf)?\s*[:\uff1a]\s*(.+)", cleaned)
        if match:
            value = match.group(1).strip()
            value = re.split(r"[;\uff1b\u3002]", value, maxsplit=1)[0].strip()
            return value
    match = re.search(r"(C\S*?\s*->\s*C\S+)", note_text)
    return match.group(1).strip() if match else ""


def clean_stem(path: Path) -> str:
    stem = path.stem
    for suffix in ("_transmission_abs2", "_transmission", "_abs2", "_T"):
        if stem.endswith(suffix):
            stem = stem[: -len(suffix)]
    return stem


def scan_index(text: str) -> str:
    match = re.match(r"^0*(\d+)(?:_|$)", text)
    return match.group(1) if match else ""


def value_nm_from_name(text: str) -> str:
    matches = re.findall(r"([-+]?\d+(?:\.\d+)?)\s*nm", text, flags=re.IGNORECASE)
    return matches[-1] if matches else ""


def first_value(row: dict[str, str], names: tuple[str, ...]) -> str:
    for name in names:
        value = row.get(name)
        if value not in (None, ""):
            return value
    return ""


def row_index(row: dict[str, str]) -> str:
    value = first_value(row, ("index", "global_index", "idx"))
    return str(int(value)) if str(value).isdigit() else str(value)


def row_params(*rows: dict[str, str]) -> dict[str, str]:
    params: dict[str, str] = {}
    for row in rows:
        for key, value in row.items():
            if key in PARAM_SKIP_KEYS or value in (None, ""):
                continue
            if key in ("index", "global_index", "idx", "name"):
                continue
            params.setdefault(key, str(value))
    return params


def primary_param_nm(params: dict[str, str], name: str) -> str:
    for key in (
        "value_nm",
        "delta_R_nm",
        "delta_nm",
        "offset_nm",
        "radius_nm",
        "long_axis_difference_nm",
        "short_axis_difference_nm",
        "height_difference_nm",
        "width_difference_nm",
    ):
        if key in params and params[key] != "":
            return params[key]
    return value_nm_from_name(name)


def existing_manifest_path(raw: str, fallback_dir: Path | None) -> Path:
    raw = (raw or "").strip()
    if raw:
        path = Path(raw)
        if path.is_file():
            return path
    if not fallback_dir:
        return Path()
    name = Path(raw).name if raw else ""
    if name:
        candidate = fallback_dir / name
        if candidate.is_file():
            return candidate
    return Path()


def fsp_for_item(name: str, manifest: dict[str, str], png_path: Path, xlsx_path: Path) -> Path:
    raw = first_value(manifest, ("fsp", "fsp_file"))
    if raw:
        path = Path(raw)
        if path.is_file() and is_under_root(path):
            return path

    run_dir = None
    for source in (xlsx_path, png_path):
        if source and source.is_file():
            run_dir = source.parent.parent
            break
    if not run_dir or not run_dir.is_dir():
        return Path()

    idx = scan_index(name)
    for dirname in FSP_DIR_NAMES:
        folder = run_dir / dirname
        if not folder.is_dir():
            continue
        direct = folder / (name + ".fsp")
        if direct.is_file():
            return direct
        if idx:
            matches = sorted(folder.glob(f"{int(idx):04d}_*.fsp"), key=lambda p: natural_key(p.name))
            if matches:
                return matches[0]
    return Path()


def add_file_keys(target: dict[str, Path], path: Path):
    base = clean_stem(path)
    idx = scan_index(base)
    target.setdefault(base, path)
    if idx:
        target.setdefault(idx, path)


def file_maps(run_dir: Path) -> tuple[dict[str, Path], dict[str, Path], Path | None, Path | None]:
    png_dir = child_dir(run_dir, PNG_DIR_NAMES)
    xlsx_dir = child_dir(run_dir, XLSX_DIR_NAMES)
    png_map: dict[str, Path] = {}
    xlsx_map: dict[str, Path] = {}
    if png_dir:
        for path in sorted(png_dir.glob("*.png"), key=lambda p: natural_key(p.name)):
            add_file_keys(png_map, path)
    if xlsx_dir:
        for path in sorted(xlsx_dir.glob("*.xlsx"), key=lambda p: natural_key(p.name)):
            add_file_keys(xlsx_map, path)
    return png_map, xlsx_map, png_dir, xlsx_dir


def norm_header(value) -> str:
    return str(value or "").strip().lower().replace(" ", "_").replace("|", "").replace("^", "")


def to_float(value) -> float | None:
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def choose_xlsx_columns(headers: list[str]) -> tuple[int | None, int | None, int | None, int | None, float]:
    normalized = [norm_header(header) for header in headers]
    x_idx = None
    x_scale = 1.0
    for idx, header in enumerate(normalized):
        if "wavelength_nm" in header or header in ("lambda_nm", "\u03bb_nm"):
            x_idx = idx
            x_scale = 1.0
            break
    if x_idx is None:
        for idx, header in enumerate(normalized):
            if "wavelength_m" in header or header in ("lambda_m", "\u03bb_m"):
                x_idx = idx
                x_scale = 1e9
                break
    if x_idx is None:
        for idx, header in enumerate(normalized):
            if "wavelength" in header or "lambda" in header:
                x_idx = idx
                x_scale = 1.0
                break
    if x_idx is None:
        x_idx = 0 if headers else None

    y_idx = None
    for idx, header in enumerate(normalized):
        if header in ("transmission_abs2", "t_abs2") or "abs2" in header:
            y_idx = idx
            break
    if y_idx is None:
        for idx, header in enumerate(normalized):
            if header in ("transmission", "transmission_raw", "t") or "transmission" in header:
                y_idx = idx
                break

    real_idx = imag_idx = None
    for idx, header in enumerate(normalized):
        if header in ("t_real", "transmission_real"):
            real_idx = idx
        if header in ("t_imag", "transmission_imag"):
            imag_idx = idx
    return x_idx, y_idx, real_idx, imag_idx, x_scale


def xlsx_data(path: Path) -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)

    def parse_sheet(sheet_name: str) -> dict:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {"columns": [], "points": [], "metrics": {}, "sheet": sheet_name}

        best_header_row = None
        best_headers: list[str] = []
        for candidate in range(min(8, len(rows))):
            headers = [str(v).strip() if v is not None else "" for v in rows[candidate]]
            normalized = [norm_header(header) for header in headers]
            has_x = any("wavelength" in header or "lambda" in header for header in normalized)
            has_y = any("abs2" in header or "transmission" in header for header in normalized)
            if has_x and has_y:
                best_header_row = candidate
                best_headers = headers
                break
        if best_header_row is None:
            return {"columns": [], "points": [], "metrics": {}, "sheet": sheet_name}

        x_idx, y_idx, real_idx, imag_idx, x_scale = choose_xlsx_columns(best_headers)
        points: list[list[float]] = []
        for row in rows[best_header_row + 1 :]:
            values = list(row)
            if x_idx is None or x_idx >= len(values):
                continue
            x = to_float(values[x_idx])
            if x is None:
                continue
            x *= x_scale

            y = None
            if y_idx is not None and y_idx < len(values):
                y = to_float(values[y_idx])
                y_header = norm_header(best_headers[y_idx] if y_idx < len(best_headers) else "")
                if y is not None and ("raw" in y_header or y_header in ("t", "transmission")):
                    y = y * y
            if y is None and real_idx is not None and real_idx < len(values):
                real = to_float(values[real_idx]) or 0.0
                imag = to_float(values[imag_idx]) if imag_idx is not None and imag_idx < len(values) else 0.0
                imag = imag or 0.0
                y = real * real + imag * imag
            if y is None:
                continue
            points.append([x, y])

        points.sort(key=lambda point: point[0])
        metrics = metrics_from_points(points)
        y_name = best_headers[y_idx] if y_idx is not None and y_idx < len(best_headers) else "Transmission_abs2"
        return {"columns": ["Wavelength_nm", y_name], "points": points, "metrics": metrics, "sheet": sheet_name}

    parsed = [parse_sheet(sheet_name) for sheet_name in wb.sheetnames]
    parsed = [data for data in parsed if data.get("points")]
    if not parsed:
        return {"columns": [], "points": [], "metrics": {}, "sheet": ""}
    return max(parsed, key=lambda data: len(data.get("points", [])))


def metrics_from_points(points: list[list[float]]) -> dict:
    if not points:
        return {}
    min_point = min(points, key=lambda point: point[1])
    max_point = max(points, key=lambda point: point[1])
    return {
        "min_x": min_point[0],
        "min_y": min_point[1],
        "max_x": max_point[0],
        "max_y": max_point[1],
        "unconverged": max_point[1] > 1.0,
    }


def summarize_xlsx(path: Path) -> dict:
    if not path.is_file():
        return {}
    data = xlsx_data(path)
    metrics = data.get("metrics", {})
    return {
        "x_column": data.get("columns", ["", ""])[0],
        "y_column": data.get("columns", ["", ""])[1] if len(data.get("columns", [])) > 1 else "",
        "min_x": metrics.get("min_x", ""),
        "min_y": metrics.get("min_y", ""),
        "max_x": metrics.get("max_x", ""),
        "max_y": metrics.get("max_y", ""),
        "unconverged": metrics.get("unconverged", False),
    }


def item_from_files(
    index: str,
    name: str,
    png_path: Path,
    xlsx_path: Path,
    scan: dict[str, str],
    manifest: dict[str, str] | None = None,
) -> dict:
    manifest = manifest or {}
    params = row_params(scan, manifest)
    value_nm = primary_param_nm(params, name)
    excel_summary = summarize_xlsx(xlsx_path) if xlsx_path.is_file() else {}
    fsp_path = fsp_for_item(name, manifest, png_path, xlsx_path)
    max_abs2 = first_value(manifest, ("max_abs2", "max_T", "max")) or str(excel_summary.get("max_y", ""))
    min_abs2 = first_value(manifest, ("min_abs2", "min_T", "min")) or str(excel_summary.get("min_y", ""))
    max_wavelength_nm = first_value(manifest, ("max_wavelength_nm", "max_lambda_nm")) or str(excel_summary.get("max_x", ""))
    min_wavelength_nm = first_value(manifest, ("min_wavelength_nm", "min_lambda_nm")) or str(excel_summary.get("min_x", ""))
    try:
        max_float = float(max_abs2)
    except (TypeError, ValueError):
        max_float = float("nan")
    unconverged = max_float > 1.0
    return {
        "index": index,
        "name": name,
        "status": first_value(manifest, ("status",)),
        "value_nm": value_nm,
        "params": params,
        "step_nm": first_value(scan, ("step_nm", "step", "step_size_nm")),
        "elapsed_s": first_value(manifest, ("elapsed_s", "time_s")),
        "max_abs2": max_abs2,
        "max_wavelength_nm": max_wavelength_nm,
        "min_abs2": min_abs2,
        "min_wavelength_nm": min_wavelength_nm,
        "unconverged": unconverged,
        "x_column": excel_summary.get("x_column", ""),
        "y_column": excel_summary.get("y_column", ""),
        "png_id": register(png_path) if png_path.is_file() else "",
        "xlsx_id": register(xlsx_path) if xlsx_path.is_file() else "",
        "fsp_id": register(fsp_path) if fsp_path.is_file() else "",
        "png_name": png_path.name if png_path.is_file() else "",
        "xlsx_name": xlsx_path.name if xlsx_path.is_file() else "",
        "fsp_name": fsp_path.name if fsp_path.is_file() else "",
    }


def summarize_run(run_dir: Path) -> dict:
    manifest_path = run_dir / "04_logs" / "manifest.csv"
    scan_plan = run_dir / "00_scan_plan" / "scan_points.csv"
    note = first_note(run_dir)

    manifest_rows = read_csv_rows(manifest_path) if manifest_path.exists() else []
    scan_rows = read_csv_rows(scan_plan) if scan_plan.exists() else []
    png_map, xlsx_map, png_dir, xlsx_dir = file_maps(run_dir)

    scan_by_key: dict[str, dict[str, str]] = {}
    for row in scan_rows:
        idx = row_index(row)
        if idx != "":
            scan_by_key[idx] = row
        if row.get("name"):
            scan_by_key.setdefault(row["name"], row)

    manifest_by_key: dict[str, dict[str, str]] = {}
    for row in manifest_rows:
        idx = row_index(row)
        if idx != "":
            manifest_by_key[idx] = row
        if row.get("name"):
            manifest_by_key[row["name"]] = row

    items: list[dict] = []
    used: set[str] = set()

    for png_path in sorted({p for p in png_map.values()}, key=lambda p: natural_key(p.name)):
        name = clean_stem(png_path)
        index = scan_index(name)
        key = index or name
        xlsx_path = xlsx_map.get(name) or xlsx_map.get(index, Path())
        scan = scan_by_key.get(index) or scan_by_key.get(name) or {}
        manifest = manifest_by_key.get(index) or manifest_by_key.get(name) or {}
        items.append(item_from_files(index, name, png_path, xlsx_path, scan, manifest))
        used.add(key)

    for xlsx_path in sorted({p for p in xlsx_map.values()}, key=lambda p: natural_key(p.name)):
        name = clean_stem(xlsx_path)
        index = scan_index(name)
        key = index or name
        if key in used:
            continue
        scan = scan_by_key.get(index) or scan_by_key.get(name) or {}
        manifest = manifest_by_key.get(index) or manifest_by_key.get(name) or {}
        items.append(item_from_files(index, name, Path(), xlsx_path, scan, manifest))
        used.add(key)

    for row in manifest_rows:
        name = row.get("name", "")
        index = row_index(row)
        key = index or name
        if key in used:
            continue
        png_path = existing_manifest_path(row.get("png", ""), png_dir)
        xlsx_path = existing_manifest_path(row.get("xlsx", ""), xlsx_dir)
        if not png_path.is_file() and not xlsx_path.is_file():
            continue
        if not name:
            if png_path.is_file():
                name = clean_stem(png_path)
            elif xlsx_path.is_file():
                name = clean_stem(xlsx_path)
            else:
                name = f"scan_{index or len(items)}"
        if not index:
            index = scan_index(name)
        scan = scan_by_key.get(index) or scan_by_key.get(name) or {}
        items.append(item_from_files(index, name, png_path, xlsx_path, scan, row))
        used.add(key)

    items.sort(key=lambda item: natural_key(str(item.get("index") or item.get("name") or "")))
    for pos, item in enumerate(items):
        label = item.get("index") or item.get("name") or str(pos)
        item["uid"] = f"{pos}:{label}"

    rel_parts = run_dir.relative_to(ROOT).parts
    label_parts = list(rel_parts)
    if "results" in label_parts:
        i = label_parts.index("results")
        group = " / ".join(label_parts[:i])
        perturbation = label_parts[i + 1] if i + 1 < len(label_parts) else ""
    else:
        group = " / ".join(label_parts[:-1])
        perturbation = ""

    note_text = read_text_guess(note, 12000) if note and note.exists() else ""

    return {
        "id": register(run_dir),
        "name": run_dir.name,
        "relative_path": to_rel(run_dir),
        "group": group,
        "perturbation": perturbation,
        "reduction_path": parse_reduction_path(note_text),
        "modified": run_dir.stat().st_mtime,
        "viewed": is_run_viewed(run_dir),
        "viewed_at": run_viewed_at(run_dir),
        "tags": run_tags(run_dir),
        "count": len(items),
        "ok_count": sum(1 for item in items if item.get("status") in ("ok", "") and not item.get("unconverged")),
        "has_manifest": manifest_path.exists(),
        "has_scan_plan": scan_plan.exists(),
        "has_note": bool(note and note.exists()),
        "manifest_id": register(manifest_path) if manifest_path.exists() else "",
        "scan_plan_id": register(scan_plan) if scan_plan.exists() else "",
        "note_id": register(note) if note and note.exists() else "",
        "note": note_text,
        "items": items,
    }


def scan_runs_uncached() -> dict:
    FILES.clear()
    runs: list[dict] = []
    if not ROOT.exists():
        return {"root": str(ROOT), "exists": False, "runs": []}

    archive_name = ARCHIVE_DIR_NAME
    candidate_dirs: set[Path] = set()
    for manifest in ROOT.rglob("04_logs/manifest.csv"):
        if archive_name in manifest.parts:
            continue
        candidate_dirs.add(manifest.parent.parent)
    for name in PNG_DIR_NAMES:
        for png_dir in ROOT.rglob(name):
            if archive_name in png_dir.parts:
                continue
            if any(png_dir.glob("*.png")):
                candidate_dirs.add(png_dir.parent)
    for name in XLSX_DIR_NAMES:
        for xlsx_dir in ROOT.rglob(name):
            if archive_name in xlsx_dir.parts:
                continue
            if any(xlsx_dir.glob("*.xlsx")):
                candidate_dirs.add(xlsx_dir.parent)

    for run_dir in sorted(candidate_dirs, key=run_sort_key, reverse=True):
        if APP_DIR in run_dir.parents:
            continue
        try:
            summary = summarize_run(run_dir)
            if summary["count"] > 0:
                runs.append(summary)
        except Exception as exc:
            runs.append(
                {
                    "id": register(run_dir),
                    "name": run_dir.name,
                    "relative_path": to_rel(run_dir),
                    "group": "scan failed",
                    "perturbation": "",
                    "modified": run_dir.stat().st_mtime if run_dir.exists() else 0,
                    "count": 0,
                    "ok_count": 0,
                    "error": str(exc),
                    "items": [],
                }
            )
    manager_payload = manager_records_payload()
    return {"root": str(ROOT), "exists": True, "runs": runs, "perturbation_status": scan_perturbation_status(runs), "archive_summary": manager_payload.get("summary", {})}


def refresh_scan_cache_background():
    global SCAN_REFRESHING, SCAN_REFRESH_STARTED, SCAN_REFRESH_FINISHED
    with SCAN_LOCK:
        if SCAN_REFRESHING:
            return
        SCAN_REFRESHING = True
        SCAN_REFRESH_STARTED = time.time()

    def worker():
        global SCAN_REFRESHING, SCAN_REFRESH_FINISHED
        try:
            payload = scan_runs_uncached()
            payload["cached"] = False
            save_scan_cache(payload)
        finally:
            with SCAN_LOCK:
                SCAN_REFRESHING = False
                SCAN_REFRESH_FINISHED = time.time()

    threading.Thread(target=worker, daemon=True).start()


def scan_runs(force: bool = False, cache_only: bool = False) -> dict:
    if not force:
        cached = load_scan_cache()
        if cached is not None:
            if not cache_only:
                refresh_scan_cache_background()
            cached["refreshing"] = SCAN_REFRESHING
            cached["refresh_started"] = SCAN_REFRESH_STARTED
            cached["refresh_finished"] = SCAN_REFRESH_FINISHED
            return cached
        if cache_only:
            return {"root": str(ROOT), "exists": ROOT.exists(), "runs": [], "perturbation_status": [], "cached": False, "empty_cache": True}
    payload = scan_runs_uncached()
    payload["cached"] = False
    payload["refreshing"] = False
    save_scan_cache(payload)
    return payload


def scan_refresh_status() -> dict:
    cached = load_scan_cache()
    return {
        "ok": True,
        "refreshing": SCAN_REFRESHING,
        "refresh_started": SCAN_REFRESH_STARTED,
        "refresh_finished": SCAN_REFRESH_FINISHED,
        "cache_time": cached.get("cache_time", 0) if cached else 0,
        "has_cache": cached is not None,
    }


def latest_full_run(folder: Path) -> Path | None:
    if not folder.exists():
        return None
    archive_name = ARCHIVE_DIR_NAME
    candidates = []
    for path in folder.rglob("run_full_*"):
        if not path.is_dir():
            continue
        if archive_name in path.parts or UNCONVERGED_FOLDER_NAME in path.parts:
            continue
        candidates.append(path)
    if not candidates:
        return None
    return max(candidates, key=lambda path: path.stat().st_mtime)


def useful_run_result_count(run_dir: Path | None) -> int:
    if not run_dir or not run_dir.exists():
        return 0
    count = 0
    for name in PNG_DIR_NAMES:
        png_dir = run_dir / name
        if png_dir.exists():
            count += len(list(png_dir.glob("*.png")))
    for name in XLSX_DIR_NAMES:
        xlsx_dir = run_dir / name
        if xlsx_dir.exists():
            count += len(list(xlsx_dir.glob("*.xlsx")))
    if (run_dir / "04_logs" / "manifest.csv").exists():
        count += 1
    return count


def max_abs2_in_run(run_dir: Path | None) -> float | None:
    if not run_dir or not run_dir.exists():
        return None
    values = []
    manifest = run_dir / "04_logs" / "manifest.csv"
    if manifest.exists():
        for row in read_csv_rows(manifest):
            normalized = {norm_header(k): v for k, v in row.items()}
            value = first_value(row, ("max_abs2", "max_t", "maxT", "max_transmission")) or first_value(normalized, ("max_abs2", "max_t", "maxt", "max_transmission"))
            try:
                if value != "":
                    values.append(float(value))
            except Exception:
                pass
    try:
        _png_map, xlsx_map, _png_dir, _xlsx_dir = file_maps(run_dir)
        for xlsx_path in xlsx_map.values():
            summary = summarize_xlsx(xlsx_path)
            value = summary.get("max_y", "")
            try:
                if value != "":
                    values.append(float(value))
            except Exception:
                pass
    except Exception:
        pass
    return max(values) if values else None


def max_abs2_from_run_summary(run: dict | None) -> float | None:
    if not run:
        return None
    values = []
    for item in run.get("items", []) or []:
        value = item.get("max_abs2")
        try:
            if value not in (None, ""):
                values.append(float(value))
        except Exception:
            pass
    return max(values) if values else None


def cached_run_summary_by_path() -> dict[str, dict]:
    cached = load_scan_cache()
    if not cached:
        return {}
    return {run.get("relative_path", ""): run for run in cached.get("runs", []) if run.get("relative_path")}


def find_script(coding_folder: Path) -> Path | None:
    scripts = sorted(coding_folder.glob("run_*.py"), key=lambda path: natural_key(path.name))
    if not scripts:
        scripts = sorted(coding_folder.glob("*.py"), key=lambda path: natural_key(path.name))
    return scripts[0] if scripts else None


def scan_perturbation_status(runs: list[dict]) -> list[dict]:
    run_by_path = {run["relative_path"]: run for run in runs}
    statuses: list[dict] = []
    skip_roots = {APP_DIR.name, ".idea", "__pycache__", "controller_logs", ARCHIVE_DIR_NAME, "\u7fa4\u8bba\u6bcd\u7ed3\u6784\u6570\u636e\u5e93_\u4e2d\u6587"}

    for category in sorted([p for p in ROOT.iterdir() if p.is_dir() and p.name not in skip_roots], key=lambda p: natural_key(p.name)):
        for structure in sorted([p for p in category.iterdir() if p.is_dir()], key=lambda p: natural_key(p.name)):
            coding_root = structure / "coding"
            if not coding_root.exists():
                continue
            for coding_folder in sorted([p for p in coding_root.iterdir() if p.is_dir()], key=lambda p: natural_key(p.name)):
                result_folder = structure / "results" / coding_folder.name
                latest_run = latest_full_run(result_folder)
                result_count = useful_run_result_count(latest_run)
                script = find_script(coding_folder)
                latest_rel = to_rel(latest_run) if latest_run else ""
                matched_run = run_by_path.get(latest_rel)
                statuses.append(
                    {
                        "id": register(coding_folder),
                        "structure": f"{category.name} / {structure.name}",
                        "category": category.name,
                        "mother": structure.name,
                        "perturbation": coding_folder.name,
                        "coding_folder": to_rel(coding_folder),
                        "result_folder": to_rel(result_folder),
                        "has_latest_run": bool(latest_run and result_count > 0),
                        "latest_run": latest_rel,
                        "latest_run_id": matched_run.get("id", "") if matched_run else "",
                        "latest_count": result_count,
                        "script_id": register(script) if script and script.is_file() else "",
                        "script_name": script.name if script else "",
                    }
                )
    return statuses


def xlsx_series(path: Path) -> dict:
    return xlsx_data(path)


def spectral_config_payload(overrides: dict | None = None) -> dict:
    config_file = SPECTRAL_CONFIG_FILE if SPECTRAL_CONFIG_FILE.exists() else LEGACY_SPECTRAL_CONFIG_FILE
    config = spectral_api.load_config(config_file)
    return spectral_api.merge_config(config, overrides or {})


def spectral_scan_payload(force: bool = False) -> dict:
    data = scan_runs(force=force, cache_only=False)
    if data.get("empty_cache") or not data.get("runs"):
        data = scan_runs(force=True)
    return data


def spectral_run_from_id(run_id: str, scan_payload: dict | None = None) -> tuple[Path | None, dict]:
    scan_payload = scan_payload or spectral_scan_payload(False)
    for run in scan_payload.get("runs", []):
        if str(run.get("id", "")) == str(run_id):
            run_dir = (ROOT / str(run.get("relative_path", ""))).resolve()
            if run_dir.is_dir() and is_under_root(run_dir):
                return run_dir, run
    ref = FILES.get(str(run_id))
    if ref and ref.path.exists():
        run_dir = ref.path if ref.path.is_dir() else run_dir_from_any_path(ref.path)
        if run_dir and run_dir.is_dir() and is_under_root(run_dir):
            return run_dir, {}
    return None, {}


def spectral_query_overrides(qs: dict) -> dict:
    overrides = {}
    target = qs.get("target", [""])[0]
    if target:
        overrides["target_type"] = target
    try:
        top_n = int(qs.get("top_n", ["0"])[0] or 0)
        if top_n > 0:
            overrides["global_top_n"] = min(500, top_n)
    except ValueError:
        pass
    return overrides


def spectral_diagnostics_response(qs: dict) -> dict:
    force = qs.get("refresh", [""])[0] in ("1", "true", "yes")
    run_id = qs.get("run_id", [""])[0] or qs.get("id", [""])[0]
    config = spectral_config_payload(spectral_query_overrides(qs))
    scan_payload = spectral_scan_payload(force=force)
    if run_id:
        run_dir, run = spectral_run_from_id(run_id, scan_payload)
        if not run_dir:
            return {"ok": False, "error": "run not found", "mode": "run"}
        return spectral_api.analyze_run_dir(run_dir, ROOT, run, config, register_file=register, include_points=True, write_outputs=True)
    return spectral_api.analyze_global(ROOT, scan_payload, config, register_file=register, force=force)


def spectral_metrics_response(qs: dict) -> dict:
    force = qs.get("refresh", [""])[0] in ("1", "true", "yes")
    run_id = qs.get("run_id", [""])[0] or qs.get("id", [""])[0]
    config = spectral_config_payload(spectral_query_overrides(qs))
    scan_payload = spectral_scan_payload(force=force)
    if run_id:
        run_dir, run = spectral_run_from_id(run_id, scan_payload)
        if not run_dir:
            return {"ok": False, "error": "run not found", "mode": "run"}
        payload = spectral_api.analyze_run_dir(run_dir, ROOT, run, config, register_file=register, include_points=False, write_outputs=True)
        return {
            "ok": payload.get("ok", False),
            "mode": "run",
            "run": payload.get("run", {}),
            "summary": payload.get("summary", {}),
            "rankings": payload.get("rankings", {}),
            "items": [{k: v for k, v in item.items() if k != "points"} for item in payload.get("items", [])],
            "missing_data": payload.get("missing_data", []),
            "export_files": payload.get("export_files", {}),
        }
    payload = spectral_api.analyze_global(ROOT, scan_payload, config, register_file=register, force=force)
    return {"ok": True, "mode": "global", "summary": payload.get("summary", {}), "rankings": payload.get("rankings", {}), "runs": payload.get("runs", [])}


def spectral_export_response(data: dict, qs: dict | None = None) -> dict:
    qs = qs or {}
    run_id = data.get("run_id") or data.get("id") or qs.get("run_id", [""])[0] or qs.get("id", [""])[0]
    force = bool(data.get("refresh")) or qs.get("refresh", [""])[0] in ("1", "true", "yes")
    config = spectral_config_payload(data.get("config") if isinstance(data.get("config"), dict) else spectral_query_overrides(qs))
    scan_payload = spectral_scan_payload(force=force)
    if run_id:
        run_dir, run = spectral_run_from_id(str(run_id), scan_payload)
        if not run_dir:
            return {"ok": False, "error": "run not found", "mode": "run"}
        payload = spectral_api.analyze_run_dir(run_dir, ROOT, run, config, register_file=register, include_points=False, write_outputs=True)
        return {"ok": True, "mode": "run", "run": payload.get("run", {}), "summary": payload.get("summary", {}), "files": payload.get("export_files", {})}
    payload = spectral_api.analyze_global(ROOT, scan_payload, config, register_file=register, force=force)
    files = spectral_api.write_global_exports(ROOT, payload, register_file=register)
    return {"ok": True, "mode": "global", "summary": payload.get("summary", {}), "files": files}


def manager_records_payload() -> dict:
    records = []
    for idx, perturbation_dir in enumerate(discover_perturbation_result_dirs_for_manager(ROOT), 1):
        run_dirs = sorted([p for p in perturbation_dir.iterdir() if p.is_dir() and p.name.lower().startswith("run_")], key=lambda p: p.stat().st_mtime, reverse=True)
        rel = perturbation_dir.relative_to(ROOT).parts if is_under_root(perturbation_dir) else perturbation_dir.parts
        if "results" in rel:
            r_idx = rel.index("results")
            category = rel[0] if r_idx >= 2 else ""
            mother = rel[r_idx - 1] if r_idx >= 1 else ""
            perturbation = rel[r_idx + 1] if r_idx + 1 < len(rel) else perturbation_dir.name
        else:
            category = mother = ""
            perturbation = perturbation_dir.name
        modes = {"full": 0, "test": 0, "preview": 0, "unknown": 0}
        for run_dir in run_dirs:
            modes[run_mode_from_name(run_dir)] = modes.get(run_mode_from_name(run_dir), 0) + 1
        archive_stats = archive_stats_for_perturbation(perturbation_dir)
        record = {
            "id": register(perturbation_dir),
            "index": idx,
            "path": to_rel(perturbation_dir),
            "category": category,
            "mother": mother,
            "perturbation": perturbation,
            "current_runs": len(run_dirs),
            "latest": run_dirs[0].name if run_dirs else "",
            "modes": modes,
            **archive_stats,
        }
        records.append(record)
    return {"ok": True, "records": records, "summary": archive_summary(records)}


def discover_perturbation_result_dirs_for_manager(root: Path) -> list[Path]:
    found = []
    for results_dir in root.rglob("results"):
        if not results_dir.is_dir() or ARCHIVE_DIR_NAME in results_dir.parts or UNCONVERGED_FOLDER_NAME in results_dir.parts:
            continue
        for child in results_dir.iterdir():
            if not child.is_dir() or child.name == ARCHIVE_DIR_NAME:
                continue
            has_run = any(p.is_dir() and p.name.lower().startswith("run_") for p in child.iterdir())
            has_old = (child / ARCHIVE_DIR_NAME).exists()
            if has_run or has_old:
                found.append(child)
    return sorted(found, key=lambda p: str(p))


def archive_stats_for_perturbation(perturbation_dir: Path) -> dict:
    old_root = perturbation_dir / ARCHIVE_DIR_NAME
    stats = {"archive_runs": 0, "archive_files": 0, "unconverged_runs": 0, "unconverged_files": 0}
    if not old_root.exists():
        return stats
    for path in old_root.rglob("*"):
        if path.is_dir() and path.name.lower().startswith("run_"):
            stats["archive_runs"] += 1
            if UNCONVERGED_QUALITY_NAME in path.parts:
                stats["unconverged_runs"] += 1
        elif path.is_file():
            stats["archive_files"] += 1
            if UNCONVERGED_QUALITY_NAME in path.parts:
                stats["unconverged_files"] += 1
    return stats


def archive_summary(records: list[dict] | None = None) -> dict:
    records = records if records is not None else manager_records_payload().get("records", [])
    return {
        "archive_runs": sum(int(r.get("archive_runs", 0)) for r in records),
        "archive_files": sum(int(r.get("archive_files", 0)) for r in records),
        "unconverged_runs": sum(int(r.get("unconverged_runs", 0)) for r in records),
        "unconverged_files": sum(int(r.get("unconverged_files", 0)) for r in records),
    }


def normalize_manager_records(record_ids: list[str] | None = None) -> dict:
    dirs = run_refs_from_ids(record_ids or []) if record_ids else discover_perturbation_result_dirs_for_manager(ROOT)
    moved = []
    skipped = []
    for perturbation_dir in dirs:
        try:
            runs = sorted([p for p in perturbation_dir.iterdir() if p.is_dir() and p.name.lower().startswith("run_")], key=lambda p: p.stat().st_mtime, reverse=True)
            for run_dir in runs[1:]:
                dest = archive_quality_destination(run_dir, "\u5f85\u8003\u5bdf")
                shutil.move(str(run_dir), str(dest))
                moved.append({"from": to_rel(run_dir), "to": to_rel(dest)})
        except Exception as exc:
            skipped.append({"path": to_rel(perturbation_dir), "reason": str(exc)})
    return {"ok": True, "moved": moved, "skipped": skipped}


def remove_children(folder: Path) -> int:
    if not folder.exists() or not folder.is_dir():
        return 0
    deleted = 0
    for child in folder.iterdir():
        if child.is_dir():
            shutil.rmtree(str(child))
        else:
            child.unlink()
        deleted += 1
    return deleted


def clean_invalid_manager_records(record_ids: list[str] | None = None) -> dict:
    dirs = run_refs_from_ids(record_ids or []) if record_ids else discover_perturbation_result_dirs_for_manager(ROOT)
    deleted = []
    skipped = []
    for perturbation_dir in dirs:
        try:
            old_root = perturbation_dir / ARCHIVE_DIR_NAME
            if not old_root.exists():
                continue
            for invalid_dir in old_root.rglob("鏃犳晥"):
                if invalid_dir.is_dir() and is_under_root(invalid_dir):
                    count = remove_children(invalid_dir)
                    if count:
                        deleted.append({"path": to_rel(invalid_dir), "count": count})
        except Exception as exc:
            skipped.append({"path": to_rel(perturbation_dir), "reason": str(exc)})
    return {"ok": True, "deleted": deleted, "skipped": skipped}


def archive_all_current_manager_records(record_ids: list[str] | None = None) -> dict:
    dirs = run_refs_from_ids(record_ids or []) if record_ids else discover_perturbation_result_dirs_for_manager(ROOT)
    moved = []
    skipped = []
    for perturbation_dir in dirs:
        try:
            runs = sorted([p for p in perturbation_dir.iterdir() if p.is_dir() and p.name.lower().startswith("run_")], key=lambda p: p.stat().st_mtime, reverse=True)
            for run_dir in runs:
                dest = archive_quality_destination(run_dir, "寰呰€冨療")
                shutil.move(str(run_dir), str(dest))
                moved.append({"from": to_rel(run_dir), "to": to_rel(dest)})
        except Exception as exc:
            skipped.append({"path": to_rel(perturbation_dir), "reason": str(exc)})
    return {"ok": True, "moved": moved, "skipped": skipped}


def delete_old_manager_records(record_ids: list[str] | None = None) -> dict:
    dirs = run_refs_from_ids(record_ids or []) if record_ids else discover_perturbation_result_dirs_for_manager(ROOT)
    deleted = []
    skipped = []
    for perturbation_dir in dirs:
        try:
            runs = sorted([p for p in perturbation_dir.iterdir() if p.is_dir() and p.name.lower().startswith("run_")], key=lambda p: p.stat().st_mtime, reverse=True)
            for run_dir in runs[KEEP_LATEST_ACTIVE_RUNS:]:
                if not is_under_root(run_dir):
                    skipped.append({"path": to_rel(run_dir), "reason": "outside root"})
                    continue
                shutil.rmtree(str(run_dir))
                deleted.append(to_rel(run_dir))
        except Exception as exc:
            skipped.append({"path": to_rel(perturbation_dir), "reason": str(exc)})
    return {"ok": True, "deleted": deleted, "skipped": skipped}


def discover_all_scripts() -> list[dict]:
    """Find all run_fdtd_*.py scripts under coding/ directories."""
    scripts = []
    seen = set()
    skip_roots = {APP_DIR.name, ".idea", "__pycache__", "controller_logs", "缇よ姣嶇粨鏋勬暟鎹簱_涓枃"}
    for category in sorted(ROOT.iterdir()):
        if not category.is_dir() or category.name in skip_roots:
            continue
        for structure in sorted(category.iterdir()):
            if not structure.is_dir():
                continue
            coding_root = structure / "coding"
            if not coding_root.exists():
                continue
            for coding_folder in sorted(coding_root.iterdir()):
                if not coding_folder.is_dir():
                    continue
                for script in sorted(coding_folder.glob("run_*.py"), key=lambda p: natural_key(p.name)):
                    rp = str(script.resolve())
                    if rp not in seen:
                        seen.add(rp)
                        result_folder = structure / "results" / coding_folder.name
                        latest_run = latest_full_run(result_folder)
                        latest_count = useful_run_result_count(latest_run)
                        scripts.append({
                            "path": script,
                            "name": f"{category.name}/{structure.name}/{coding_folder.name}/{script.name}",
                            "category": category.name,
                            "structure": structure.name,
                            "perturbation": coding_folder.name,
                            "result_folder": to_rel(result_folder),
                            "has_latest_run": bool(latest_run and latest_count > 0),
                            "latest_run": to_rel(latest_run) if latest_run else "",
                            "latest_count": latest_count,
                        })
    scripts.sort(key=lambda x: (x["category"], x["structure"], x["perturbation"], Path(x["path"]).name))
    for idx, script in enumerate(scripts, 1):
        script["id"] = idx
    return scripts


def controller_script_payload() -> list[dict]:
    payload = []
    cached_runs = cached_run_summary_by_path()
    for item in discover_all_scripts():
        params = script_param_groups(item["path"])
        result_folder = item["path"].parent.parent.parent / "results" / item["perturbation"]
        latest_run = latest_full_run(result_folder)
        latest_count = useful_run_result_count(latest_run)
        latest_rel = to_rel(latest_run) if latest_run else ""
        latest_max_abs2 = max_abs2_from_run_summary(cached_runs.get(latest_rel))
        if latest_max_abs2 is None:
            latest_max_abs2 = max_abs2_in_run(latest_run)
        payload.append({
            "id": item.get("id"),
            "name": item.get("name"),
            "category": item.get("category"),
            "structure": item.get("structure"),
            "perturbation": item.get("perturbation"),
            "script": str(item.get("path")),
            "params": params,
            "result_folder": to_rel(result_folder),
            "has_latest_run": bool(latest_run and latest_count > 0),
            "latest_run": latest_rel,
            "latest_count": latest_count,
            "latest_max_abs2": latest_max_abs2,
            "unconverged": bool(latest_max_abs2 is not None and latest_max_abs2 > 1.0),
        })
    return payload


def filter_scripts_by_ids(all_scripts: list[dict], ids_str: str) -> list[dict]:
    """Filter scripts by comma-separated 1-based indices like '1,3,5-8'."""
    indices = set()
    for part in ids_str.replace("，", ",").split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            a, b = part.split("-", 1)
            try:
                start, end = int(a), int(b)
                if start > end:
                    start, end = end, start
                indices.update(range(start, end + 1))
            except ValueError:
                pass
        else:
            try:
                indices.add(int(part))
            except ValueError:
                pass
    return [all_scripts[i - 1] for i in sorted(indices) if 1 <= i <= len(all_scripts)]


def start_parallel_scripts_job(cmds: list[dict]) -> str:
    """Run multiple scripts in parallel, return a single job_id."""
    job_id = uuid.uuid4().hex[:12]
    job = job_template(job_id, "parallel_scripts", [], ROOT)
    job["total"] = len(cmds)
    job["lines"] = [f"prepare to run {len(cmds)} scripts in parallel..."]
    with JOBS_LOCK:
        JOBS[job_id] = job

    def worker():
        procs = []
        try:
            env = os.environ.copy()
            env["PYTHONIOENCODING"] = "utf-8:replace"
            env["PYTHONUNBUFFERED"] = "1"
            for entry in cmds:
                proc = subprocess.Popen(entry["cmd"], cwd=entry["cwd"], stdout=subprocess.PIPE,
                                        stderr=subprocess.STDOUT, stdin=subprocess.DEVNULL, text=True,
                                        encoding="utf-8", errors="replace", bufsize=1, env=env)
                procs.append((proc, entry))
                with JOBS_LOCK:
                    JOB_PROCS[job_id] = proc
                    job["pid"] = proc.pid
                    job["current"] = min(len(procs), len(cmds))
                    job["current_title"] = entry["title"]
                    job["progress"] = round(job["current"] * 100 / max(1, job["total"]))
                    job["lines"].append(f"started {entry['title']} (PID {proc.pid})")
            for proc, entry in procs:
                for line in proc.stdout or []:
                    line = line.rstrip("\r\n")
                    if useful_output_line(line):
                        with JOBS_LOCK:
                            job["lines"].append(f"[{entry['title']}] {line}")
                            job["lines"] = job["lines"][-20000:]
                            update_job_from_line(job, line)
                proc.wait()
                with JOBS_LOCK:
                    status = "done" if proc.returncode == 0 else f"failed (code={proc.returncode})"
                    job["lines"].append(f"{entry['title']}: {status}")
            with JOBS_LOCK:
                failures = sum(1 for p, _ in procs if p.returncode != 0)
                job["returncode"] = 0 if failures == 0 else -1
                job["progress"] = 100
                job["lines"].append(f"all done. success {len(procs) - failures}, failed {failures}.")
        except Exception as exc:
            with JOBS_LOCK:
                job["lines"].append(f"ERROR: {exc}")
                job["returncode"] = -1
        finally:
            with JOBS_LOCK:
                JOB_PROCS.pop(job_id, None)
                job["running"] = False
            for entry in cmds:
                for path in entry.get("temps", []):
                    try:
                        if path.exists() and path.name.startswith("_web_temp_"):
                            path.unlink()
                    except Exception:
                        pass

    threading.Thread(target=worker, daemon=True).start()
    return job_id


def start_sequential_scripts_job(cmds: list[dict]) -> str:
    """Run scripts one by one, return a single job_id."""
    job_id = uuid.uuid4().hex[:12]
    job = job_template(job_id, "sequential_scripts", [], ROOT)
    job["total"] = len(cmds)
    job["lines"] = [f"prepare to run {len(cmds)} scripts sequentially..."]
    with JOBS_LOCK:
        JOBS[job_id] = job

    def worker():
        success = 0
        fail = 0
        try:
            env = os.environ.copy()
            env["PYTHONIOENCODING"] = "utf-8:replace"
            env["PYTHONUNBUFFERED"] = "1"
            for idx, entry in enumerate(cmds, 1):
                with JOBS_LOCK:
                    job["current"] = idx
                    job["current_title"] = entry["title"]
                    job["progress"] = round(idx * 100 / max(1, len(cmds)))
                    job["lines"].append(f"\n{'='*60}")
                    job["lines"].append(f"[{idx}/{len(cmds)}] start: {entry['title']}")
                try:
                    proc = subprocess.Popen(entry["cmd"], cwd=entry["cwd"], stdout=subprocess.PIPE,
                                            stderr=subprocess.STDOUT, stdin=subprocess.DEVNULL, text=True,
                                            encoding="utf-8", errors="replace", bufsize=1, env=env)
                    with JOBS_LOCK:
                        JOB_PROCS[job_id] = proc
                        job["pid"] = proc.pid
                    for line in proc.stdout or []:
                        line = line.rstrip("\r\n")
                        if useful_output_line(line):
                            with JOBS_LOCK:
                                job["lines"].append(line)
                                job["lines"] = job["lines"][-20000:]
                                update_job_from_line(job, line)
                    proc.wait()
                    if proc.returncode == 0:
                        success += 1
                        with JOBS_LOCK:
                            job["lines"].append(f"[{entry['title']}] done")
                    else:
                        fail += 1
                        with JOBS_LOCK:
                            job["lines"].append(f"[{entry['title']}] failed (code={proc.returncode})")
                except Exception as exc:
                    fail += 1
                    with JOBS_LOCK:
                        job["lines"].append(f"[{entry['title']}] exception: {exc}")
                finally:
                    with JOBS_LOCK:
                        JOB_PROCS.pop(job_id, None)
                    for path in entry.get("temps", []):
                        try:
                            if path.exists() and path.name.startswith("_web_temp_"):
                                path.unlink()
                        except Exception:
                            pass
            with JOBS_LOCK:
                job["returncode"] = 0 if fail == 0 else -1
                job["progress"] = 100
                job["lines"].append(f"\n{'='*60}")
                job["lines"].append(f"all done. success {success}, failed {fail}, total {len(cmds)}.")
        except Exception as exc:
            with JOBS_LOCK:
                job["lines"].append(f"ERROR: {exc}")
                job["returncode"] = -1
        finally:
            with JOBS_LOCK:
                JOB_PROCS.pop(job_id, None)
                job["running"] = False

    threading.Thread(target=worker, daemon=True).start()
    return job_id


class Handler(BaseHTTPRequestHandler):
    def send_json(self, data, status=200):
        body = json.dumps(data, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        parsed = urlparse(self.path)
        path = unquote(parsed.path)
        qs = parse_qs(parsed.query)

        if path in ("/", "/index.html"):
            self.serve_file(APP_DIR / "index.html")
            return
        if path == "/favicon.ico":
            self.send_response(204)
            self.end_headers()
            return
        if path.startswith("/assets/"):
            asset = (APP_DIR / path.lstrip("/")).resolve()
            try:
                asset.relative_to(APP_DIR.resolve())
            except ValueError:
                self.send_error(403, "forbidden")
                return
            if not asset.is_file():
                self.send_error(404, "asset not found")
                return
            self.serve_file(asset)
            return
        if path == "/topology_transition_analysis.html":
            self.serve_file(APP_DIR / "topology_transition_analysis.html")
            return
        if path == "/spectral_physics_diagnostics.html":
            self.serve_file(APP_DIR / "spectral_physics_diagnostics.html")
            return
        spectral_suite_pages = {
            "/spectral_compare.html": "spectral_compare.html",
            "/field_phase_poynting_viewer.html": "field_phase_poynting_viewer.html",
            "/global_spectral_leaderboard.html": "global_spectral_leaderboard.html",
            "/missing_data_repair_center.html": "missing_data_repair_center.html",
            "/batch_recompute_center.html": "batch_recompute_center.html",
            "/report_preview_print.html": "report_preview_print.html",
            "/compare_set_manager.html": "compare_set_manager.html",
            "/data_quality_audit.html": "data_quality_audit.html",
            "/resource_file_browser.html": "resource_file_browser.html",
        }
        if path in spectral_suite_pages:
            self.serve_file(APP_DIR / spectral_suite_pages[path])
            return
        if path == "/api/spectral-config":
            self.send_json({"ok": True, "config": spectral_config_payload()})
            return
        if path == "/api/spectral-diagnostics":
            try:
                self.send_json(spectral_diagnostics_response(qs))
            except Exception as exc:
                self.send_json({"ok": False, "error": str(exc), "mode": "spectral"}, status=500)
            return
        if path == "/api/spectrum-metrics":
            try:
                self.send_json(spectral_metrics_response(qs))
            except Exception as exc:
                self.send_json({"ok": False, "error": str(exc), "mode": "spectral"}, status=500)
            return
        if path == "/api/spectral-export":
            try:
                self.send_json(spectral_export_response({}, qs))
            except Exception as exc:
                self.send_json({"ok": False, "error": str(exc), "mode": "spectral"}, status=500)
            return
        if path == "/api/scan":
            force = qs.get("refresh", [""])[0] in ("1", "true", "yes")
            cache_only = qs.get("cache_only", [""])[0] in ("1", "true", "yes")
            self.send_json(scan_runs(force=force, cache_only=cache_only))
            return
        if path == "/api/scan-status":
            self.send_json(scan_refresh_status())
            return
        if path == "/api/file":
            fid = qs.get("id", [""])[0]
            ref = FILES.get(fid)
            if not ref or not ref.path.exists():
                self.send_error(404, "file not found")
                return
            self.serve_file(ref.path)
            return
        if path == "/api/text":
            fid = qs.get("id", [""])[0]
            ref = FILES.get(fid)
            if not ref or not ref.path.exists():
                self.send_error(404, "file not found")
                return
            self.send_json({"text": read_text_guess(ref.path, 200000)})
            return
        if path == "/api/xlsx":
            fid = qs.get("id", [""])[0]
            ref = FILES.get(fid)
            if not ref or not ref.path.exists():
                self.send_error(404, "file not found")
                return
            try:
                self.send_json(xlsx_series(ref.path))
            except Exception as exc:
                self.send_json({"error": str(exc), "columns": [], "points": []}, status=500)
            return
        if path == "/api/job":
            jid = qs.get("id", [""])[0]
            try:
                since = max(0, int(qs.get("since", ["0"])[0] or 0))
            except ValueError:
                since = 0
            with JOBS_LOCK:
                job = JOBS.get(jid)
                if job:
                    refresh_job_latest_spectrum(job)
                payload = dict(job) if job else None
                if payload:
                    lines = list(job.get("lines", []))
                    if since > len(lines):
                        since = max(0, len(lines) - 500)
                    payload["line_count"] = len(lines)
                    payload["lines"] = lines[since:] if since else lines
            if not payload:
                self.send_json({"ok": False, "error": "job not found"}, status=404)
                return
            payload["ok"] = True
            self.send_json(payload)
            return
        if path == "/api/jobs":
            with JOBS_LOCK:
                jobs = sorted((job_summary(job) for job in JOBS.values()), key=lambda item: float(item.get("started") or 0), reverse=True)
            active = [job for job in jobs if job.get("running")]
            latest = active[0] if active else (jobs[0] if jobs else None)
            self.send_json({"ok": True, "jobs": jobs[:20], "latest": latest})
            return
        if path == "/api/script-params":
            fid = qs.get("id", [""])[0]
            ref = FILES.get(fid)
            if not ref or not ref.path.is_file() or ref.path.suffix.lower() != ".py":
                self.send_json({"ok": False, "error": "script not found"}, status=404)
                return
            self.send_json({"ok": True, "script": str(ref.path), "params": script_param_groups(ref.path)})
            return
        if path == "/api/controller-params":
            controller = ROOT / "fdtd_master_controller.py"
            scripts = controller_script_payload()
            if controller.exists():
                self.send_json({"ok": True, "controller": str(controller), "params": script_param_groups(controller), "scripts": scripts})
            else:
                self.send_json({"ok": True, "controller": "", "params": {"scan": [], "runtime": {}, "values": {}}, "scripts": scripts, "note": "no controller found; will run individual scripts"})
            return
        if path == "/api/results-manager":
            self.send_json(manager_records_payload())
            return
        if path == "/api/open-folder":
            fid = qs.get("id", [""])[0]
            ref = FILES.get(fid)
            if not ref or not ref.path.exists():
                self.send_json({"ok": False, "error": "folder not found"}, status=404)
                return
            target = ref.path if ref.path.is_dir() else ref.path.parent
            try:
                if os.name == "nt":
                    subprocess.Popen(["explorer", str(target.resolve())])
                else:
                    os.startfile(str(target))
                self.send_json({"ok": True, "path": str(target)})
            except Exception as exc:
                self.send_json({"ok": False, "error": str(exc)}, status=500)
            return
        if path == "/api/open-file":
            fid = qs.get("id", [""])[0]
            ref = FILES.get(fid)
            if not ref or not ref.path.exists() or not ref.path.is_file() or not is_under_root(ref.path):
                self.send_json({"ok": False, "error": "file not found"}, status=404)
                return
            try:
                if os.name == "nt":
                    os.startfile(str(ref.path.resolve()))
                else:
                    subprocess.Popen(["xdg-open", str(ref.path.resolve())])
                self.send_json({"ok": True, "path": str(ref.path)})
            except Exception as exc:
                self.send_json({"ok": False, "error": str(exc)}, status=500)
            return

        self.send_error(404, "not found")

    def do_POST(self):
        parsed = urlparse(self.path)
        path = unquote(parsed.path)
        qs = parse_qs(parsed.query)

        if path == "/api/spectral-config":
            try:
                data = read_json_body(self)
                config = spectral_api.save_config(SPECTRAL_CONFIG_FILE, data.get("config", data))
                self.send_json({"ok": True, "config": config})
            except Exception as exc:
                self.send_json({"ok": False, "error": str(exc)}, status=500)
            return
        if path == "/api/spectral-export":
            try:
                data = read_json_body(self)
                self.send_json(spectral_export_response(data, qs))
            except Exception as exc:
                self.send_json({"ok": False, "error": str(exc), "mode": "spectral"}, status=500)
            return

        if path == "/api/stop-job":
            try:
                data = read_json_body(self)
                jid = data.get("job_id") or data.get("id") or qs.get("id", [""])[0]
                with JOBS_LOCK:
                    job = JOBS.get(jid)
                    proc = JOB_PROCS.get(jid)
                    pid = proc.pid if proc else (job or {}).get("pid")
                    if job:
                        job["stop_requested"] = True
                        job["current_detail"] = "stopping..."
                        job["lines"].append("stop requested, terminating controller and child processes.")
                if not job:
                    self.send_json({"ok": False, "error": "job not found"}, status=404)
                    return
                kill_process_tree(int(pid or 0))
                with JOBS_LOCK:
                    job["running"] = False
                    job["returncode"] = -9
                    job["lines"].append("stop command sent.")
                self.send_json({"ok": True, "job_id": jid})
            except Exception as exc:
                self.send_json({"ok": False, "error": str(exc)}, status=500)
            return

        if path == "/api/mark-viewed":
            fid = qs.get("id", [""])[0]
            ref = FILES.get(fid)
            if not ref or not ref.path.exists():
                self.send_json({"ok": False, "error": "run not found"}, status=404)
                return
            target = ref.path if ref.path.is_dir() else ref.path.parent
            try:
                mark_run_viewed(target)
                self.send_json({"ok": True, "path": str(target)})
            except Exception as exc:
                self.send_json({"ok": False, "error": str(exc)}, status=500)
            return

        if path == "/api/run-tags":
            try:
                data = read_json_body(self)
                fid = str(data.get("id") or "")
                ref = FILES.get(fid)
                if not ref or not ref.path.exists():
                    self.send_json({"ok": False, "error": "run not found"}, status=404)
                    return
                target = ref.path if ref.path.is_dir() else ref.path.parent
                tags = data.get("tags") or []
                if not isinstance(tags, list):
                    tags = []
                saved = set_run_tags(target, tags)
                self.send_json({"ok": True, "id": fid, "tags": saved})
            except Exception as exc:
                self.send_json({"ok": False, "error": str(exc)}, status=500)
            return

        if path == "/api/batch-run-tags":
            try:
                data = read_json_body(self)
                ids = data.get("run_ids") or []
                tags = data.get("tags") or []
                mode = str(data.get("mode") or "add")
                if not isinstance(ids, list):
                    ids = []
                if not isinstance(tags, list):
                    tags = []
                updated = []
                for fid in ids:
                    ref = FILES.get(str(fid))
                    if not ref or not ref.path.exists():
                        continue
                    target = ref.path if ref.path.is_dir() else ref.path.parent
                    current = run_tags(target)
                    if mode == "set":
                        next_tags = [str(tag) for tag in tags]
                    elif mode == "remove":
                        remove = {str(tag) for tag in tags}
                        next_tags = [tag for tag in current if tag not in remove]
                    else:
                        next_tags = current + [str(tag) for tag in tags]
                    updated.append({"id": str(fid), "tags": set_run_tags(target, next_tags)})
                self.send_json({"ok": True, "updated": updated})
            except Exception as exc:
                self.send_json({"ok": False, "error": str(exc)}, status=500)
            return

        if path == "/api/move-unconverged":
            try:
                data = read_json_body(self)
                run_ids = data.get("run_ids") or []
                job_id = start_operation_job("move_unconverged_runs", lambda: move_unconverged_runs(run_ids))
                self.send_json({"ok": True, "job_id": job_id})
            except Exception as exc:
                self.send_json({"ok": False, "error": str(exc)}, status=500)
            return

        if path == "/api/delete-item":
            try:
                data = read_json_body(self)
                deleted = []
                for key in ("png_id", "xlsx_id"):
                    fid = data.get(key, "")
                    ref = FILES.get(fid)
                    if ref and ref.path.exists() and ref.path.is_file() and is_under_root(ref.path):
                        recycle_path(ref.path)
                        deleted.append(str(ref.path))
                self.send_json({"ok": True, "deleted": deleted})
            except Exception as exc:
                self.send_json({"ok": False, "error": str(exc)}, status=500)
            return

        if path == "/api/delete-runs-by-view-state":
            try:
                data = read_json_body(self)
                want_viewed = bool(data.get("viewed"))
                run_ids = data.get("run_ids") or []

                def delete_action():
                    deleted = []
                    skipped = []
                    run_dirs = run_refs_from_ids(run_ids)
                    if not run_dirs:
                        scan = scan_runs()
                        for run in scan.get("runs", []):
                            if bool(run.get("viewed")) != want_viewed:
                                continue
                            ref = FILES.get(run.get("id", ""))
                            if ref and ref.path.is_dir():
                                run_dirs.append(ref.path)
                    for run_dir in run_dirs:
                        try:
                            recycle_folder(run_dir)
                            deleted.append(to_rel(run_dir))
                        except PermissionError as exc:
                            skipped.append({"path": to_rel(run_dir), "reason": f"file in use: {exc}"})
                        except OSError as exc:
                            skipped.append({"path": to_rel(run_dir), "reason": str(exc)})
                        except RuntimeError as exc:
                            skipped.append({"path": to_rel(run_dir), "reason": str(exc)})
                    return {"ok": True, "deleted": deleted, "skipped": skipped}

                job_id = start_operation_job("delete_runs", delete_action)
                self.send_json({"ok": True, "job_id": job_id})
            except Exception as exc:
                self.send_json({"ok": False, "error": str(exc)}, status=500)
            return

        if path == "/api/start-controller":
            try:
                data = read_json_body(self)
                mode = data.get("mode") or "full"
                style = data.get("style") or "sequential"
                missing_only = bool(data.get("missing_only"))
                max_parallel = int(data.get("max_parallel") or 1)
                child_timeout_s = float(data.get("child_timeout_s") or 3600)
                ids_str = str(data.get("ids") or "").strip()
                raw_overrides = data.get("overrides") or {}
                overrides = normalize_override_payload(raw_overrides)

                controller = ROOT / "fdtd_master_controller.py"
                if controller.exists():
                    override_file = None
                    cmd = [sys.executable, str(controller), "--mode", mode, "--style", style, "--max-parallel", str(max(1, max_parallel)), "--child-timeout-s", str(max(0, child_timeout_s)), "--yes"]
                    if ids_str:
                        cmd += ["--ids", ids_str]
                    else:
                        cmd += ["--all"]
                    if missing_only:
                        cmd += ["--missing-only"]
                    if overrides:
                        override_dir = GENERATED_DIR / "controller_overrides"
                        override_dir.mkdir(parents=True, exist_ok=True)
                        fd, name = tempfile.mkstemp(prefix="fdtd_controller_overrides_", suffix=".json", dir=str(override_dir))
                        os.close(fd)
                        override_file = Path(name)
                        override_file.write_text(json.dumps(overrides, ensure_ascii=False), encoding="utf-8")
                        cmd += ["--overrides-json", str(override_file)]
                    job_id = start_background_job(cmd, ROOT, "fdtd_master_controller", [override_file] if override_file else [])
                    self.send_json({"ok": True, "job_id": job_id})
                else:
                    # Fallback: run individual sweep scripts directly
                    all_scripts = discover_all_scripts()
                    if not all_scripts:
                        self.send_json({"ok": False, "error": "no sweep scripts found"}, status=404)
                        return
                    selected = filter_scripts_by_ids(all_scripts, ids_str) if ids_str else all_scripts
                    if missing_only:
                        selected = [script_info for script_info in selected if not script_info.get("has_latest_run")]
                    if not selected:
                        self.send_json({"ok": False, "error": f"ids {ids_str} did not match any scripts"}, status=404)
                        return
                    if style == "parallel":
                        cmds = []
                        for script_info in selected:
                            script_overrides = overrides.get(str(script_info["path"])) or overrides.get("*") or {}
                            temp = temporary_script(script_info["path"], script_overrides) if script_overrides else script_info["path"]
                            cmds.append({
                                "cmd": command_for_script(temp, mode),
                                "cwd": str(script_info["path"].parent),
                                "title": script_info["name"],
                                "temps": [temp] if temp != script_info["path"] else [],
                            })
                        job_id = start_parallel_scripts_job(cmds)
                    else:
                        cmds = []
                        for script_info in selected:
                            script_overrides = overrides.get(str(script_info["path"])) or overrides.get("*") or {}
                            temp = temporary_script(script_info["path"], script_overrides) if script_overrides else script_info["path"]
                            cmds.append({
                                "cmd": command_for_script(temp, mode),
                                "cwd": str(script_info["path"].parent),
                                "title": script_info["name"],
                                "temps": [temp] if temp != script_info["path"] else [],
                            })
                        job_id = start_sequential_scripts_job(cmds)
                    self.send_json({"ok": True, "job_id": job_id, "script_count": len(selected)})
            except Exception as exc:
                self.send_json({"ok": False, "error": str(exc)}, status=500)
            return

        if path == "/api/results-manager-normalize":
            try:
                data = read_json_body(self)
                ids = data.get("ids") or []
                job_id = start_operation_job("results_manager_normalize", lambda: normalize_manager_records(ids))
                self.send_json({"ok": True, "job_id": job_id})
            except Exception as exc:
                self.send_json({"ok": False, "error": str(exc)}, status=500)
            return

        if path == "/api/results-manager-clean-invalid":
            try:
                data = read_json_body(self)
                ids = data.get("ids") or []
                job_id = start_operation_job("results_manager_clean_invalid", lambda: clean_invalid_manager_records(ids))
                self.send_json({"ok": True, "job_id": job_id})
            except Exception as exc:
                self.send_json({"ok": False, "error": str(exc)}, status=500)
            return

        if path == "/api/results-manager-archive-current":
            try:
                data = read_json_body(self)
                ids = data.get("ids") or []
                job_id = start_operation_job("results_manager_archive_current", lambda: archive_all_current_manager_records(ids))
                self.send_json({"ok": True, "job_id": job_id})
            except Exception as exc:
                self.send_json({"ok": False, "error": str(exc)}, status=500)
            return

        if path == "/api/results-manager-delete-old":
            try:
                data = read_json_body(self)
                ids = data.get("ids") or []
                job_id = start_operation_job("results_manager_delete_old", lambda: delete_old_manager_records(ids))
                self.send_json({"ok": True, "job_id": job_id})
            except Exception as exc:
                self.send_json({"ok": False, "error": str(exc)}, status=500)
            return

        if path == "/api/delete-run":
            fid = qs.get("id", [""])[0]
            ref = FILES.get(fid)
            if not ref or not ref.path.exists():
                self.send_json({"ok": False, "error": "run folder not found"}, status=404)
                return
            try:
                recycle_folder(ref.path)
                self.send_json({"ok": True, "path": str(ref.path)})
            except Exception as exc:
                self.send_json({"ok": False, "error": str(exc)}, status=500)
            return

        if path == "/api/start-script":
            fid = qs.get("id", [""])[0]
            ref = FILES.get(fid)
            if not ref or not ref.path.is_file():
                self.send_json({"ok": False, "error": "script not found"}, status=404)
                return
            if not is_under_root(ref.path) or ref.path.suffix.lower() != ".py":
                self.send_json({"ok": False, "error": "refusing to run this file"}, status=400)
                return
            try:
                data = read_json_body(self)
                overrides = data.get("overrides") or {}
                mode = data.get("mode") or "full"
                script = temporary_script(ref.path, overrides)
                cmd = command_for_script(script, mode)
                job_id = start_background_job(cmd, script.parent, ref.path.name, [script] if script != ref.path else [])
                self.send_json({"ok": True, "job_id": job_id, "script": str(ref.path)})
            except Exception as exc:
                self.send_json({"ok": False, "error": str(exc)}, status=500)
            return

        self.send_error(404, "not found")

    def serve_file(self, path: Path):
        if not path.exists() or path.is_dir():
            self.send_error(404, "file not found")
            return
        ctype = mimetypes.guess_type(str(path))[0] or "application/octet-stream"
        if ctype.startswith("text/") or path.suffix.lower() in (".html", ".js", ".css", ".md"):
            ctype = f"{ctype}; charset=utf-8"
        data = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def log_message(self, fmt, *args):
        print("%s - %s" % (self.address_string(), fmt % args))


def main():
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 8787
    server = ThreadingHTTPServer(("127.0.0.1", port), Handler)
    url = f"http://127.0.0.1:{port}/"
    print(f"FDTD result viewer: {url}")
    print(f"Scanning root: {ROOT}")
    try:
        webbrowser.open(url)
    except Exception:
        pass
    server.serve_forever()


if __name__ == "__main__":
    main()
