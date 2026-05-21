from __future__ import annotations

import csv
import json
import math
import re
import statistics
import time
from pathlib import Path
from typing import Any, Callable


SPECTRUM_EXTS = {".xlsx", ".csv", ".txt"}
IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".webp", ".svg"}
DATA_EXTS = SPECTRUM_EXTS | IMAGE_EXTS | {".json", ".npy", ".npz", ".mat", ".h5", ".hdf5"}

TARGET_TYPES = (
    "notch",
    "passband",
    "fano",
    "q_mode",
    "edge",
    "broadband_high",
    "broadband_low",
    "flat",
    "custom",
)

DIR_PROFILES = {
    "transmission": {
        "labels": ("T", "透射谱"),
        "dirs": ("02_transmission_excel", "03_transmission_excel", "transmission", "transmission_excel"),
        "image_dirs": ("03_transmission_abs2_png", "03_transmission_png_abs2", "transmission_png"),
        "value_aliases": (
            "t",
            "transmission",
            "transmittance",
            "t_abs2",
            "transmission_abs2",
            "abs2",
            "|t|2",
            "|t|^2",
            "t2",
            "power",
            "透射",
            "透射率",
        ),
    },
    "reflection": {
        "labels": ("R", "反射谱"),
        "dirs": ("06_reflection_excel", "reflection", "reflection_excel"),
        "image_dirs": ("06_reflection_png", "reflection_png"),
        "value_aliases": ("r", "reflection", "reflectance", "r_abs2", "|r|2", "|r|^2", "反射", "反射率"),
    },
    "absorption": {
        "labels": ("A", "吸收谱"),
        "dirs": ("07_absorption_excel", "absorption", "loss", "absorption_excel"),
        "image_dirs": ("07_absorption_png", "absorption_png"),
        "value_aliases": ("a", "absorption", "absorptance", "loss", "absorbed", "吸收", "损耗", "吸收率"),
    },
}

SUPPORT_DIRS = {
    "field": {
        "folder": "08_field_data",
        "label": "场图",
        "why": "场图用于判断模式是否局域、是否位于高折射率区域或边界附近。",
        "next": "在候选峰/谷中心波长和背景波长处导出 Ex/Ey/Ez、Hx/Hy/Hz 与 |E|^2。",
    },
    "phase": {
        "folder": "09_phase_data",
        "label": "相位数据",
        "why": "相位有助于识别 Fano 干涉、相位跃迁、涡旋或拓扑相关特征。",
        "next": "在共振附近导出相位谱和关键平面的相位分布。",
    },
    "poynting": {
        "folder": "10_poynting_data",
        "label": "Poynting 能流",
        "why": "Poynting 矢量可用于判断能流阻断、泄露方向和局域能量循环。",
        "next": "在候选共振波长处导出 Sx/Sy/Sz 或矢量场图。",
    },
}

WAVELENGTH_ALIASES = (
    "wavelength",
    "wavelength_nm",
    "lambda",
    "lambda_nm",
    "lambda_m",
    "lambda_um",
    "wl",
    "wl_nm",
    "x",
    "波长",
)

PARAM_SKIP_KEYS = {
    "status",
    "fsp",
    "fsp_file",
    "xlsx",
    "excel_file",
    "csv",
    "png",
    "png_file",
    "elapsed_s",
    "max_abs2",
    "max_wavelength_nm",
    "min_abs2",
    "min_wavelength_nm",
    "name",
}


def default_config() -> dict[str, Any]:
    return {
        "target_type": "auto",
        "max_points_per_spectrum": 1200,
        "global_top_n": 80,
        "quality": {
            "min_points": 20,
            "warn_points": 40,
            "max_nan_ratio": 0.05,
            "soft_upper": 1.05,
            "hard_upper": 1.2,
            "negative_tolerance": -0.02,
            "edge_fraction": 0.05,
            "flat_dynamic_threshold": 0.02,
        },
        "custom_weights": {
            "peak": 1.0,
            "dip": 1.0,
            "q": 1.0,
            "background": 1.0,
            "flatness": 0.4,
            "quality": 1.0,
        },
    }


def merge_config(base: dict[str, Any] | None, incoming: dict[str, Any] | None) -> dict[str, Any]:
    out = json.loads(json.dumps(base or default_config(), ensure_ascii=False))
    if not isinstance(incoming, dict):
        return out
    for key, value in incoming.items():
        if isinstance(value, dict) and isinstance(out.get(key), dict):
            out[key].update(value)
        else:
            out[key] = value
    return out


def load_config(path: Path) -> dict[str, Any]:
    if not path.exists():
        return default_config()
    try:
        return merge_config(default_config(), json.loads(path.read_text(encoding="utf-8")))
    except Exception:
        return default_config()


def save_config(path: Path, config: dict[str, Any]) -> dict[str, Any]:
    merged = merge_config(default_config(), config)
    path.write_text(json.dumps(merged, ensure_ascii=False, indent=2), encoding="utf-8")
    return merged


def is_under_root(path: Path, root: Path) -> bool:
    try:
        path.resolve().relative_to(root.resolve())
        return True
    except ValueError:
        return False


def rel_path(path: Path, root: Path) -> str:
    try:
        return str(path.resolve().relative_to(root.resolve()))
    except ValueError:
        return str(path)


def read_text_guess(path: Path, limit: int | None = None) -> str:
    data = path.read_bytes()
    if limit:
        data = data[:limit]
    for enc in ("utf-8-sig", "utf-8", "gb18030", "cp936"):
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    return data.decode("utf-8", errors="replace")


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    try:
        text = read_text_guess(path)
        return list(csv.DictReader(text.splitlines()))
    except Exception:
        return []


def normalize_name(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("λ", "lambda").replace("μ", "u").replace("µ", "u")
    text = text.replace("²", "2").replace("^", "").replace("|", "")
    text = re.sub(r"[\s\-\(\)\[\]\{\}/\\]+", "_", text)
    text = text.strip("_")
    return text


def clean_stem(path: Path | str) -> str:
    name = Path(str(path)).stem
    return re.sub(r"(_transmission_abs2|_abs2|_transmission|_reflection|_absorption)$", "", name, flags=re.I)


def scan_index(text: str) -> str:
    match = re.search(r"(?<!\d)(\d{1,6})(?!\d)", str(text))
    return str(int(match.group(1))) if match else ""


def natural_key(value: Any) -> list[Any]:
    parts = re.split(r"(\d+)", str(value))
    return [int(part) if part.isdigit() else part.lower() for part in parts]


def to_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if math.isfinite(float(value)):
            return float(value)
        return None
    text = str(value).strip().replace(",", "")
    if not text or text.lower() in {"nan", "inf", "-inf", "none"}:
        return None
    try:
        out = float(text)
        return out if math.isfinite(out) else None
    except ValueError:
        return None


def clipped(value: float, lo: float = 0.0, hi: float = 1.0) -> float:
    if not math.isfinite(value):
        return lo
    return min(hi, max(lo, value))


def safe_mean(values: list[float]) -> float | None:
    clean = [float(v) for v in values if math.isfinite(float(v))]
    return statistics.fmean(clean) if clean else None


def percentile(values: list[float], q: float) -> float | None:
    clean = sorted(float(v) for v in values if math.isfinite(float(v)))
    if not clean:
        return None
    if len(clean) == 1:
        return clean[0]
    pos = clipped(q, 0, 1) * (len(clean) - 1)
    lo = int(math.floor(pos))
    hi = int(math.ceil(pos))
    if lo == hi:
        return clean[lo]
    return clean[lo] + (clean[hi] - clean[lo]) * (pos - lo)


def choose_column(headers: list[Any], kind: str) -> tuple[int | None, int | None, str, str]:
    normalized = [normalize_name(h) for h in headers]
    x_idx = None
    y_idx = None
    x_name = ""
    y_name = ""
    for idx, header in enumerate(normalized):
        raw = str(headers[idx] or "")
        if any(alias == header or alias in header for alias in WAVELENGTH_ALIASES):
            x_idx = idx
            x_name = raw
            break
    if x_idx is None and headers:
        x_idx = 0
        x_name = str(headers[0] or "")

    aliases = DIR_PROFILES.get(kind, DIR_PROFILES["transmission"])["value_aliases"]
    normalized_aliases = [normalize_name(alias) for alias in aliases]
    for idx, header in enumerate(normalized):
        if idx == x_idx:
            continue
        if any(alias == header or alias in header for alias in normalized_aliases):
            y_idx = idx
            y_name = str(headers[idx] or "")
            break
    if y_idx is None:
        for idx, header in enumerate(normalized):
            if idx == x_idx:
                continue
            if kind == "transmission" and ("abs2" in header or "trans" in header):
                y_idx = idx
                y_name = str(headers[idx] or "")
                break
            if kind == "reflection" and ("reflect" in header or header == "r"):
                y_idx = idx
                y_name = str(headers[idx] or "")
                break
            if kind == "absorption" and ("absorb" in header or "loss" in header or header == "a"):
                y_idx = idx
                y_name = str(headers[idx] or "")
                break
    return x_idx, y_idx, x_name, y_name


def infer_wavelength_scale(values: list[float], column_name: str = "") -> tuple[float, str]:
    header = normalize_name(column_name)
    if "nm" in header or "纳米" in header:
        return 1.0, "nm"
    if "_m" in header or header.endswith("m") and "um" not in header:
        return 1e9, "m"
    if "um" in header or "micron" in header or "微米" in header:
        return 1000.0, "um"
    clean = [abs(v) for v in values if math.isfinite(v) and v != 0]
    if not clean:
        return 1.0, "unknown"
    med = statistics.median(clean)
    if med < 1e-3:
        return 1e9, "m_inferred"
    if med < 50:
        return 1000.0, "um_inferred"
    return 1.0, "nm_inferred"


def numeric_columns_from_rows(rows: list[list[Any]], start: int) -> tuple[int | None, int | None]:
    counts: dict[int, int] = {}
    for row in rows[start:]:
        for idx, value in enumerate(row):
            if to_float(value) is not None:
                counts[idx] = counts.get(idx, 0) + 1
    if len(counts) < 2:
        return None, None
    ordered = sorted(counts, key=lambda idx: counts[idx], reverse=True)
    return ordered[0], ordered[1]


def parse_table_rows(rows: list[list[Any]], kind: str, source: Path, sheet: str = "") -> dict[str, Any]:
    total_rows = len(rows)
    best: dict[str, Any] | None = None
    for header_row in range(min(12, len(rows))):
        headers = list(rows[header_row])
        x_idx, y_idx, x_name, y_name = choose_column(headers, kind)
        if x_idx is None or y_idx is None:
            continue
        raw_x: list[float] = []
        raw_pairs: list[tuple[float, float]] = []
        valid_rows = 0
        for row in rows[header_row + 1 :]:
            if x_idx >= len(row) or y_idx >= len(row):
                continue
            x = to_float(row[x_idx])
            y = to_float(row[y_idx])
            if x is None:
                continue
            raw_x.append(x)
            if y is None:
                continue
            raw_pairs.append((x, y))
            valid_rows += 1
        if valid_rows >= 4:
            scale, unit = infer_wavelength_scale(raw_x, x_name)
            points = sorted([[x * scale, y] for x, y in raw_pairs], key=lambda pair: pair[0])
            best = {
                "ok": True,
                "source": str(source),
                "sheet": sheet,
                "kind": kind,
                "columns": {"wavelength": x_name or str(headers[x_idx]), "value": y_name or str(headers[y_idx])},
                "unit": unit,
                "points": points,
                "total_rows": max(total_rows - header_row - 1, len(points)),
            }
            break

    if best:
        return best

    x_idx, y_idx = numeric_columns_from_rows(rows, 0)
    if x_idx is None or y_idx is None:
        return {"ok": False, "error": "未找到可识别的波长列和谱值列", "points": [], "total_rows": total_rows}
    raw_x = []
    points = []
    for row in rows:
        if x_idx >= len(row) or y_idx >= len(row):
            continue
        x = to_float(row[x_idx])
        y = to_float(row[y_idx])
        if x is None or y is None:
            continue
        raw_x.append(x)
        points.append([x, y])
    scale, unit = infer_wavelength_scale(raw_x, "")
    return {
        "ok": bool(points),
        "source": str(source),
        "sheet": sheet,
        "kind": kind,
        "columns": {"wavelength": f"column_{x_idx + 1}", "value": f"column_{y_idx + 1}"},
        "unit": unit,
        "points": sorted([[x * scale, y] for x, y in points], key=lambda pair: pair[0]),
        "total_rows": total_rows,
        "warning": "未找到表头，已按前两个数值列解析",
    }


def read_csv_spectrum(path: Path, kind: str) -> dict[str, Any]:
    text = read_text_guess(path)
    sample = "\n".join(text.splitlines()[:10])
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t ")
    except Exception:
        dialect = csv.excel
    rows = list(csv.reader(text.splitlines(), dialect))
    return parse_table_rows(rows, kind, path)


def read_xlsx_spectrum(path: Path, kind: str) -> dict[str, Any]:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    parsed = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = [list(row) for row in ws.iter_rows(values_only=True)]
        data = parse_table_rows(rows, kind, path, sheet_name)
        if data.get("points"):
            parsed.append(data)
    if not parsed:
        return {"ok": False, "error": "Excel 中没有可解析谱线", "points": [], "total_rows": 0}
    return max(parsed, key=lambda item: len(item.get("points") or []))


def read_spectrum_file(path: Path, kind: str = "transmission") -> dict[str, Any]:
    suffix = path.suffix.lower()
    try:
        if suffix == ".xlsx":
            return read_xlsx_spectrum(path, kind)
        if suffix in {".csv", ".txt"}:
            return read_csv_spectrum(path, kind)
        return {"ok": False, "error": f"暂不支持 {suffix} 谱线解析", "points": [], "total_rows": 0}
    except Exception as exc:
        return {"ok": False, "error": str(exc), "points": [], "total_rows": 0}


def crossing_x(x1: float, y1: float, x2: float, y2: float, level: float) -> float:
    if y1 == y2:
        return (x1 + x2) / 2.0
    ratio = (level - y1) / (y2 - y1)
    return x1 + (x2 - x1) * clipped(ratio, 0.0, 1.0)


def compute_fwhm(x: list[float], y: list[float], center_idx: int, feature: str, background: float) -> dict[str, Any]:
    n = len(x)
    if n < 4 or center_idx <= 0 or center_idx >= n - 1:
        return {"value": None, "reliable": False, "reason": "特征点位于边界或采样点过少"}
    center_y = y[center_idx]
    if feature == "peak":
        prominence = center_y - background
        if prominence <= 0:
            return {"value": None, "reliable": False, "reason": "峰值相对背景不突出"}
        level = background + prominence / 2.0
        crossed = lambda yy: yy <= level
    else:
        prominence = background - center_y
        if prominence <= 0:
            return {"value": None, "reliable": False, "reason": "谷值相对背景不突出"}
        level = background - prominence / 2.0
        crossed = lambda yy: yy >= level

    left = None
    for i in range(center_idx - 1, -1, -1):
        if crossed(y[i]):
            left = crossing_x(x[i], y[i], x[i + 1], y[i + 1], level)
            break
    right = None
    for i in range(center_idx + 1, n):
        if crossed(y[i]):
            right = crossing_x(x[i - 1], y[i - 1], x[i], y[i], level)
            break

    if left is None or right is None or right <= left:
        return {"value": None, "reliable": False, "level": level, "reason": "半高/半深交点不完整，可能贴近波段边界"}
    width = right - left
    samples = sum(1 for xx in x if left <= xx <= right)
    reliable = samples >= 4 and width > 0
    return {
        "value": width,
        "reliable": reliable,
        "level": level,
        "left_nm": left,
        "right_nm": right,
        "samples": samples,
        "reason": "" if reliable else "FWHM 由过少采样点决定，建议加密波长采样",
    }


def local_extrema(y: list[float], dynamic_range: float, noise_level: float) -> tuple[list[int], list[int]]:
    if len(y) < 3:
        return [], []
    median_y = statistics.median(y)
    threshold = max(dynamic_range * 0.08, noise_level * dynamic_range * 2.5, 1e-12)
    peaks = []
    dips = []
    for i in range(1, len(y) - 1):
        if y[i] >= y[i - 1] and y[i] >= y[i + 1] and (y[i] - median_y) >= threshold:
            peaks.append(i)
        if y[i] <= y[i - 1] and y[i] <= y[i + 1] and (median_y - y[i]) >= threshold:
            dips.append(i)
    return peaks, dips


def estimate_noise(y: list[float], dynamic_range: float) -> float:
    if len(y) < 5 or dynamic_range <= 0:
        return 0.0
    second = [abs(y[i + 1] - 2 * y[i] + y[i - 1]) for i in range(1, len(y) - 1)]
    med = percentile(second, 0.5) or 0.0
    return clipped(med / max(dynamic_range, 1e-12), 0.0, 1.0)


def metric_quality_flags(metrics: dict[str, Any], config: dict[str, Any]) -> tuple[list[dict[str, str]], float]:
    qconf = config.get("quality", {})
    flags: list[dict[str, str]] = []
    score = 1.0

    def add(code: str, level: str, message: str, penalty: float):
        nonlocal score
        flags.append({"code": code, "level": level, "message": message})
        score -= penalty

    point_count = int(metrics.get("point_count") or 0)
    if point_count < int(qconf.get("min_points", 20)):
        add("too_few_points", "error", "有效采样点过少，谱线指标不可靠", 0.55)
    elif point_count < int(qconf.get("warn_points", 40)):
        add("low_sampling", "warn", "采样点偏少，窄线宽可能被低估或高估", 0.18)
    if float(metrics.get("nan_ratio") or 0) > float(qconf.get("max_nan_ratio", 0.05)):
        add("nan_ratio_high", "error", "NaN/Inf 或不可解析点比例偏高", 0.35)
    if float(metrics.get("high_outlier_ratio") or 0) > 0.02:
        add("t_gt_1", "warn", "存在 T/R/A 超过 1 的异常点", 0.18)
    if float(metrics.get("severe_outlier_ratio") or 0) > 0.005:
        add("severe_t_gt_1", "error", "存在明显超过 1.2 的异常点，疑似归一化或收敛问题", 0.38)
    if float(metrics.get("negative_ratio") or 0) > 0.01:
        add("negative_values", "warn", "存在负谱值，建议检查监视器归一化或导出过程", 0.16)
    if float(metrics.get("dynamic_range") or 0) < float(qconf.get("flat_dynamic_threshold", 0.02)):
        add("flat_spectrum", "info", "谱线动态范围很小，不能仅凭 Q 值判断高价值", 0.18)
    if metrics.get("edge_feature"):
        add("feature_near_boundary", "warn", "主峰或主谷贴近波段边界，FWHM 可能不完整", 0.22)
    if metrics.get("sampling_insufficient"):
        add("sampling_insufficient", "warn", "FWHM 或局部特征由过少点决定", 0.22)
    if metrics.get("fwhm_reliable") is False:
        add("fwhm_unreliable", "warn", "FWHM 不可可靠计算", 0.22)
    return flags, clipped(score, 0.05, 1.0)


def extract_spectrum_features(points: list[list[float]], total_rows: int, config: dict[str, Any]) -> dict[str, Any]:
    clean = [(float(x), float(y)) for x, y in points if math.isfinite(float(x)) and math.isfinite(float(y))]
    clean = sorted(clean, key=lambda item: item[0])
    dedup: list[tuple[float, float]] = []
    for x, y in clean:
        if dedup and abs(x - dedup[-1][0]) < 1e-12:
            dedup[-1] = (x, (dedup[-1][1] + y) / 2.0)
        else:
            dedup.append((x, y))
    clean = dedup
    total = max(total_rows, len(points), len(clean))
    if not clean:
        return {
            "point_count": 0,
            "nan_ratio": 1.0,
            "quality_flags": [{"code": "no_data", "level": "error", "message": "没有可解析谱线点"}],
            "quality_score": 0.0,
        }

    x = [item[0] for item in clean]
    y = [item[1] for item in clean]
    n = len(y)
    y_min = min(y)
    y_max = max(y)
    min_idx = y.index(y_min)
    max_idx = y.index(y_max)
    dynamic = y_max - y_min
    edge_n = max(2, min(max(2, n // 8), n // 2))
    edge_values = y[:edge_n] + y[-edge_n:]
    background = safe_mean(edge_values)
    if background is None:
        background = statistics.median(y)
    global_mean = statistics.fmean(y)
    global_median = statistics.median(y)
    std = statistics.pstdev(y) if len(y) > 1 else 0.0
    noise = estimate_noise(y, dynamic)

    fwhm_peak = compute_fwhm(x, y, max_idx, "peak", background)
    fwhm_dip = compute_fwhm(x, y, min_idx, "dip", background)
    peak_q = x[max_idx] / fwhm_peak["value"] if fwhm_peak.get("value") else None
    dip_q = x[min_idx] / fwhm_dip["value"] if fwhm_dip.get("value") else None
    fwhm_candidates = [item for item in (fwhm_peak, fwhm_dip) if item.get("value")]
    fwhm_reliable = any(bool(item.get("reliable")) for item in fwhm_candidates) if fwhm_candidates else False
    q_candidates = [q for q in (peak_q, dip_q) if q and math.isfinite(q)]
    q_best = max(q_candidates) if q_candidates else None
    center_lambda = x[min_idx] if (background - y_min) >= (y_max - background) else x[max_idx]
    line_width = min((item.get("value") for item in fwhm_candidates if item.get("value")), default=None)

    peaks, dips = local_extrema(y, dynamic, noise)
    side_peak_count = len([idx for idx in peaks if idx != max_idx])
    side_dip_count = len([idx for idx in dips if idx != min_idx])

    span = max(x[-1] - x[0], 1e-12)
    edge_fraction = float(config.get("quality", {}).get("edge_fraction", 0.05))
    edge_feature = (max_idx <= n * edge_fraction or max_idx >= n * (1 - edge_fraction) - 1 or min_idx <= n * edge_fraction or min_idx >= n * (1 - edge_fraction) - 1)

    peak_prominence = max(0.0, y_max - background)
    dip_prominence = max(0.0, background - y_min)
    peak_contrast = peak_prominence / max(abs(background), 1e-12)
    dip_contrast = dip_prominence / max(abs(background), 1e-12)
    peak_dip_distance = abs(x[max_idx] - x[min_idx])
    closeness = 1.0 - clipped(peak_dip_distance / max(span * 0.25, 1e-12))

    def half_widths(fwhm: dict[str, Any], center: float) -> tuple[float | None, float | None]:
        if not fwhm.get("left_nm") or not fwhm.get("right_nm"):
            return None, None
        return center - float(fwhm["left_nm"]), float(fwhm["right_nm"]) - center

    peak_left, peak_right = half_widths(fwhm_peak, x[max_idx])
    dip_left, dip_right = half_widths(fwhm_dip, x[min_idx])
    ratios = []
    for left, right in ((peak_left, peak_right), (dip_left, dip_right)):
        if left and right and left > 0 and right > 0:
            ratios.append(abs(math.log(max(left, right) / max(min(left, right), 1e-12))))
    asymmetry = clipped((max(ratios) if ratios else 0.0) / math.log(5.0), 0.0, 1.0)

    slopes = []
    for i in range(1, n):
        dx = x[i] - x[i - 1]
        if dx:
            slopes.append(abs((y[i] - y[i - 1]) / dx))
    max_slope = max(slopes) if slopes else 0.0
    normalized_slope = clipped(max_slope * span / max(dynamic, 1e-12), 0.0, 10.0) / 10.0

    high_outlier_ratio = sum(1 for yy in y if yy > float(config.get("quality", {}).get("soft_upper", 1.05))) / max(n, 1)
    severe_outlier_ratio = sum(1 for yy in y if yy > float(config.get("quality", {}).get("hard_upper", 1.2))) / max(n, 1)
    negative_ratio = sum(1 for yy in y if yy < float(config.get("quality", {}).get("negative_tolerance", -0.02))) / max(n, 1)
    fwhm_samples = max(int(fwhm_peak.get("samples") or 0), int(fwhm_dip.get("samples") or 0))
    sampling_insufficient = n < int(config.get("quality", {}).get("warn_points", 40)) or (fwhm_candidates and fwhm_samples < 4)

    metrics: dict[str, Any] = {
        "point_count": n,
        "total_rows": total,
        "nan_ratio": clipped((total - n) / max(total, 1), 0.0, 1.0),
        "wavelength_min_nm": x[0],
        "wavelength_max_nm": x[-1],
        "wavelength_span_nm": span,
        "t_max": y_max,
        "t_min": y_min,
        "lambda_at_t_max_nm": x[max_idx],
        "lambda_at_t_min_nm": x[min_idx],
        "main_peak_lambda_nm": x[max_idx],
        "main_peak_value": y_max,
        "main_dip_lambda_nm": x[min_idx],
        "main_dip_value": y_min,
        "center_lambda_nm": center_lambda,
        "background_mean": background,
        "mean_value": global_mean,
        "median_value": global_median,
        "std_value": std,
        "ripple": std,
        "dynamic_range": dynamic,
        "peak_prominence": peak_prominence,
        "dip_prominence": dip_prominence,
        "peak_contrast": peak_contrast,
        "dip_contrast": dip_contrast,
        "fwhm_peak_nm": fwhm_peak.get("value"),
        "fwhm_dip_nm": fwhm_dip.get("value"),
        "fwhm_peak": fwhm_peak,
        "fwhm_dip": fwhm_dip,
        "line_width_nm": line_width,
        "q_peak": peak_q,
        "q_dip": dip_q,
        "q": q_best,
        "num_peaks": len(peaks),
        "num_dips": len(dips),
        "side_peak_count": side_peak_count,
        "side_dip_count": side_dip_count,
        "asymmetry": asymmetry,
        "peak_dip_distance_nm": peak_dip_distance,
        "peak_dip_closeness": closeness,
        "edge_steepness": normalized_slope,
        "max_slope": max_slope,
        "noise_level": noise,
        "edge_feature": edge_feature,
        "fwhm_reliable": fwhm_reliable,
        "sampling_insufficient": bool(sampling_insufficient),
        "high_outlier_ratio": high_outlier_ratio,
        "severe_outlier_ratio": severe_outlier_ratio,
        "negative_ratio": negative_ratio,
        "possible_nonconvergence": severe_outlier_ratio > 0 or high_outlier_ratio > 0.03,
    }
    flags, quality_score = metric_quality_flags(metrics, config)
    metrics["quality_flags"] = flags
    metrics["quality_score"] = quality_score
    return metrics


def q_norm(q: float | None) -> float:
    if not q or not math.isfinite(float(q)):
        return 0.0
    return clipped(math.log10(max(float(q), 1.0)) / 3.0, 0.0, 1.0)


def compute_scores(metrics: dict[str, Any], config: dict[str, Any]) -> dict[str, float]:
    quality = float(metrics.get("quality_score") or 0.0)
    dynamic = float(metrics.get("dynamic_range") or 0.0)
    bg = float(metrics.get("background_mean") or 0.0)
    mean_v = float(metrics.get("mean_value") or 0.0)
    std = float(metrics.get("std_value") or 0.0)
    peak = float(metrics.get("t_max") or 0.0)
    dip = float(metrics.get("t_min") or 0.0)
    peak_prom = float(metrics.get("peak_prominence") or 0.0)
    dip_prom = float(metrics.get("dip_prominence") or 0.0)
    noise = float(metrics.get("noise_level") or 0.0)
    edge_penalty = 0.25 if metrics.get("edge_feature") else 0.0
    fwhm_penalty = 0.22 if metrics.get("fwhm_reliable") is False else 0.0
    side_peak_penalty = clipped(float(metrics.get("side_peak_count") or 0) / 4.0)
    side_dip_penalty = clipped(float(metrics.get("side_dip_count") or 0) / 4.0)
    q_value = q_norm(metrics.get("q"))
    q_peak_value = q_norm(metrics.get("q_peak"))
    q_dip_value = q_norm(metrics.get("q_dip"))

    background_high = clipped(bg)
    background_low = clipped(1.0 - bg)
    depth = clipped(bg - dip)
    peak_height = clipped(peak)
    peak_contrast = clipped(peak_prom)
    dip_contrast = clipped(dip_prom)
    asym = float(metrics.get("asymmetry") or 0.0)
    close = float(metrics.get("peak_dip_closeness") or 0.0)
    edge = float(metrics.get("edge_steepness") or 0.0)
    ripple_penalty = clipped(std / max(abs(mean_v), 0.08))
    flatness = clipped(1.0 - clipped(dynamic / 0.1))

    notch_raw = (
        0.28 * depth
        + 0.20 * background_high
        + 0.22 * q_dip_value
        + 0.12 * clipped(dip_contrast)
        + 0.08 * clipped(dynamic)
        - 0.12 * side_dip_penalty
        - 0.08 * noise
        - edge_penalty
        - fwhm_penalty * 0.5
    )
    passband_raw = (
        0.28 * peak_height
        + 0.24 * peak_contrast
        + 0.18 * q_peak_value
        + 0.12 * background_low
        + 0.08 * clipped(dynamic)
        - 0.12 * side_peak_penalty
        - 0.08 * noise
        - edge_penalty
        - fwhm_penalty * 0.5
    )
    fano_raw = (
        0.24 * clipped(peak_prom + dip_prom)
        + 0.24 * asym
        + 0.18 * close
        + 0.12 * q_value
        + 0.10 * clipped(dynamic)
        - 0.10 * noise
        - edge_penalty * 0.6
    )
    q_raw = (
        0.42 * q_value
        + 0.25 * clipped(max(peak_prom, dip_prom))
        + 0.15 * clipped(dynamic)
        - 0.12 * noise
        - edge_penalty
        - fwhm_penalty
    )
    edge_raw = (
        0.38 * edge
        + 0.24 * clipped(dynamic)
        + 0.12 * clipped(abs((safe_mean([metrics.get("t_max", 0), metrics.get("t_min", 0)]) or 0) - mean_v) + dynamic)
        - 0.08 * noise
    )
    broadband_high_raw = clipped(mean_v) - 0.45 * ripple_penalty - 0.15 * noise
    broadband_low_raw = clipped(1.0 - mean_v) - 0.45 * ripple_penalty - 0.15 * noise
    flat_raw = flatness - 0.35 * noise - 0.15 * abs(mean_v - 0.5)

    weights = config.get("custom_weights", {})
    custom_raw = (
        float(weights.get("peak", 1.0)) * peak_height
        + float(weights.get("dip", 1.0)) * depth
        + float(weights.get("q", 1.0)) * q_value
        + float(weights.get("background", 1.0)) * background_high
        + float(weights.get("flatness", 0.4)) * flatness
    ) / max(sum(abs(float(v)) for v in weights.values()) or 1.0, 1.0)

    raw_scores = {
        "notch": notch_raw,
        "passband": passband_raw,
        "fano": fano_raw,
        "q_mode": q_raw,
        "edge": edge_raw,
        "broadband_high": broadband_high_raw,
        "broadband_low": broadband_low_raw,
        "flat": flat_raw,
        "custom": custom_raw,
    }
    scores = {key: round(100.0 * quality * clipped(value), 3) for key, value in raw_scores.items()}
    scores["overall"] = round(max(scores.get(k, 0.0) for k in TARGET_TYPES), 3) if scores else 0.0
    return scores


def recommendation(metrics: dict[str, Any], scores: dict[str, float]) -> dict[str, Any]:
    ordered = sorted(((key, scores.get(key, 0.0)) for key in TARGET_TYPES if key != "custom"), key=lambda item: item[1], reverse=True)
    if not ordered:
        return {"target": "custom", "confidence": 0.0, "reasons": ["没有足够数据推荐目标类型"]}
    top, top_score = ordered[0]
    second = ordered[1][1] if len(ordered) > 1 else 0.0
    confidence = clipped((top_score / 100.0) * 0.72 + clipped((top_score - second) / 35.0) * 0.28)
    reasons = []
    if top == "notch":
        reasons.append("背景透射较高且存在相对突出的低透射谷")
        if metrics.get("q_dip"):
            reasons.append("主谷线宽可估计，可用于窄带陷波比较")
    elif top == "passband":
        reasons.append("背景较低或局部峰值相对突出")
        if metrics.get("q_peak"):
            reasons.append("主峰具备可比较的线宽/Q 指标")
    elif top == "fano":
        reasons.append("峰谷相邻且线形非对称性较明显")
    elif top == "q_mode":
        reasons.append("窄线宽与足够 prominence 同时出现")
    elif top == "edge":
        reasons.append("谱线存在快速跃迁或较大斜率")
    elif top == "broadband_high":
        reasons.append("平均透射较高且波动相对可控")
    elif top == "broadband_low":
        reasons.append("平均透射较低且波动相对可控")
    elif top == "flat":
        reasons.append("谱线整体较平坦")
    if metrics.get("quality_flags"):
        warn = [flag["message"] for flag in metrics["quality_flags"] if flag.get("level") in {"warn", "error"}]
        if warn:
            reasons.append("推荐需谨慎：" + "；".join(warn[:2]))
    return {"target": top, "confidence": round(confidence, 3), "reasons": reasons}


def score_spectrum(metrics: dict[str, Any], config: dict[str, Any]) -> tuple[dict[str, float], dict[str, Any]]:
    scores = compute_scores(metrics, config)
    return scores, recommendation(metrics, scores)


def sample_points(points: list[list[float]], max_points: int) -> list[list[float]]:
    if max_points <= 0 or len(points) <= max_points:
        return [[round(float(x), 9), round(float(y), 9)] for x, y in points]
    step = len(points) / max_points
    out = []
    for i in range(max_points):
        idx = min(len(points) - 1, int(round(i * step)))
        x, y = points[idx]
        out.append([round(float(x), 9), round(float(y), 9)])
    return out


def find_child_dir(run_dir: Path, names: tuple[str, ...]) -> Path | None:
    for name in names:
        path = run_dir / name
        if path.is_dir():
            return path
    lower = {name.lower() for name in names}
    for path in run_dir.iterdir() if run_dir.exists() else []:
        if path.is_dir() and path.name.lower() in lower:
            return path
    return None


def add_file_key(mapping: dict[str, Path], path: Path) -> None:
    stem = clean_stem(path)
    idx = scan_index(stem)
    mapping.setdefault(stem, path)
    if idx:
        mapping.setdefault(idx, path)


def discover_kind_files(run_dir: Path, kind: str) -> tuple[list[Path], dict[str, Path]]:
    profile = DIR_PROFILES[kind]
    files: list[Path] = []
    image_map: dict[str, Path] = {}
    for dirname in profile["dirs"]:
        folder = run_dir / dirname
        if folder.is_dir():
            for path in sorted(folder.iterdir(), key=lambda p: natural_key(p.name)):
                if path.is_file() and path.suffix.lower() in SPECTRUM_EXTS:
                    files.append(path)
    for dirname in profile.get("image_dirs", ()):
        folder = run_dir / dirname
        if folder.is_dir():
            for path in sorted(folder.iterdir(), key=lambda p: natural_key(p.name)):
                if path.is_file() and path.suffix.lower() in IMAGE_EXTS:
                    add_file_key(image_map, path)
    return sorted(set(files), key=lambda p: natural_key(p.name)), image_map


def data_availability(run_dir: Path, root: Path, register_file: Callable[[Path], str] | None = None) -> dict[str, Any]:
    availability: dict[str, Any] = {}
    for kind, profile in DIR_PROFILES.items():
        files, images = discover_kind_files(run_dir, kind)
        availability[kind] = {
            "present": bool(files),
            "count": len(files),
            "image_count": len(images),
            "files": [{"name": p.name, "path": rel_path(p, root), "id": register_file(p) if register_file else ""} for p in files[:20]],
        }
    for kind, meta in SUPPORT_DIRS.items():
        folder = run_dir / meta["folder"]
        files = []
        if folder.is_dir():
            files = [p for p in folder.rglob("*") if p.is_file() and p.suffix.lower() in DATA_EXTS]
        availability[kind] = {
            "present": bool(files),
            "count": len(files),
            "files": [{"name": p.name, "path": rel_path(p, root), "id": register_file(p) if register_file else ""} for p in files[:30]],
        }
    return availability


def missing_data_report(availability: dict[str, Any]) -> list[dict[str, Any]]:
    report = []
    required = [
        ("reflection", "反射谱", "R 谱可帮助区分反射型共振、吸收型共振和透射归一化异常。", "导出 06_reflection_excel，并尽量保持与 T 谱相同波长采样。"),
        ("absorption", "吸收/损耗谱", "A 谱可判断低透射来自吸收损耗还是反射/带隙。", "导出 07_absorption_excel，或在有 T/R 后计算 A=1-T-R。"),
    ]
    for key, label, why, next_step in required:
        if not availability.get(key, {}).get("present"):
            report.append({"key": key, "label": label, "severity": "missing", "why": why, "next": next_step})
    for key, meta in SUPPORT_DIRS.items():
        if not availability.get(key, {}).get("present"):
            report.append({"key": key, "label": meta["label"], "severity": "missing", "why": meta["why"], "next": meta["next"]})
    if not availability.get("transmission", {}).get("present"):
        report.insert(0, {"key": "transmission", "label": "透射谱", "severity": "critical", "why": "T 谱是当前评分和谱线诊断的基础。", "next": "导出 02_transmission_excel 或兼容 CSV。"})
    return report


def row_index(row: dict[str, str]) -> str:
    for key in ("index", "idx", "scan_index", "sample_index", "i"):
        if row.get(key) not in (None, ""):
            return scan_index(str(row.get(key))) or str(row.get(key))
    for key in ("name", "sample", "png", "xlsx", "csv"):
        if row.get(key):
            idx = scan_index(str(row.get(key)))
            if idx:
                return idx
    return ""


def row_params(*rows: dict[str, str]) -> dict[str, str]:
    params: dict[str, str] = {}
    for row in rows:
        for key, value in (row or {}).items():
            if value in (None, ""):
                continue
            clean = str(key).strip()
            if normalize_name(clean) in PARAM_SKIP_KEYS:
                continue
            params[clean] = str(value)
    return params


def load_run_rows(run_dir: Path) -> tuple[dict[str, dict[str, str]], dict[str, dict[str, str]]]:
    scan_rows = read_csv_rows(run_dir / "00_scan_plan" / "scan_points.csv")
    manifest_rows = read_csv_rows(run_dir / "04_logs" / "manifest.csv")
    scan_map: dict[str, dict[str, str]] = {}
    manifest_map: dict[str, dict[str, str]] = {}
    for row in scan_rows:
        idx = row_index(row)
        if idx:
            scan_map[idx] = row
        if row.get("name"):
            scan_map.setdefault(str(row["name"]), row)
    for row in manifest_rows:
        idx = row_index(row)
        if idx:
            manifest_map[idx] = row
        if row.get("name"):
            manifest_map.setdefault(str(row["name"]), row)
    return scan_map, manifest_map


def infer_scan_axis(items: list[dict[str, Any]]) -> dict[str, Any]:
    candidates: dict[str, list[float]] = {}
    for item in items:
        for key, value in (item.get("params") or {}).items():
            number = to_float(value)
            if number is None:
                continue
            candidates.setdefault(key, []).append(number)
    ranked = []
    for key, values in candidates.items():
        if len(values) < 2:
            continue
        spread = max(values) - min(values)
        unique = len({round(v, 12) for v in values})
        if spread > 0 and unique > 1:
            ranked.append((unique, spread, key))
    if ranked:
        ranked.sort(reverse=True)
        axis = ranked[0][2]
        for item in items:
            item["scan_axis"] = axis
            item["scan_value"] = to_float((item.get("params") or {}).get(axis))
        return {"name": axis, "unit": infer_unit_from_name(axis), "source": "params"}
    for pos, item in enumerate(items):
        item["scan_axis"] = "index"
        item["scan_value"] = to_float(item.get("index")) if item.get("index") else pos
    return {"name": "index", "unit": "", "source": "fallback"}


def infer_unit_from_name(name: str) -> str:
    low = normalize_name(name)
    if "nm" in low:
        return "nm"
    if "um" in low:
        return "um"
    if "deg" in low or "angle" in low or "角" in name:
        return "deg"
    return ""


def item_png_for(path: Path, image_map: dict[str, Path]) -> Path | None:
    stem = clean_stem(path)
    idx = scan_index(stem)
    return image_map.get(stem) or image_map.get(idx or "")


def analyze_spectrum_item(
    path: Path,
    kind: str,
    run_dir: Path,
    root: Path,
    scan_map: dict[str, dict[str, str]],
    manifest_map: dict[str, dict[str, str]],
    image_map: dict[str, Path],
    config: dict[str, Any],
    register_file: Callable[[Path], str] | None = None,
    include_points: bool = True,
) -> dict[str, Any]:
    parsed = read_spectrum_file(path, kind)
    stem = clean_stem(path)
    idx = scan_index(stem)
    scan_row = scan_map.get(idx) or scan_map.get(stem) or {}
    manifest_row = manifest_map.get(idx) or manifest_map.get(stem) or {}
    params = row_params(scan_row, manifest_row)
    image = item_png_for(path, image_map)
    item: dict[str, Any] = {
        "uid": idx or stem,
        "index": idx,
        "name": stem,
        "kind": kind,
        "file_name": path.name,
        "file_path": rel_path(path, root),
        "file_id": register_file(path) if register_file else "",
        "png_name": image.name if image else "",
        "png_path": rel_path(image, root) if image else "",
        "png_id": register_file(image) if (image and register_file) else "",
        "params": params,
        "parse": {key: parsed.get(key) for key in ("ok", "error", "warning", "sheet", "columns", "unit", "total_rows")},
    }
    if not parsed.get("ok") or not parsed.get("points"):
        item["metrics"] = {
            "point_count": 0,
            "quality_score": 0.0,
            "quality_flags": [{"code": "parse_failed", "level": "error", "message": parsed.get("error") or "谱线解析失败"}],
        }
        item["scores"] = {key: 0.0 for key in (*TARGET_TYPES, "overall")}
        item["recommendation"] = {"target": "custom", "confidence": 0.0, "reasons": ["谱线解析失败，不能推荐目标类型"]}
        item["points"] = []
        return item

    points = parsed.get("points") or []
    metrics = extract_spectrum_features(points, int(parsed.get("total_rows") or len(points)), config)
    scores, rec = score_spectrum(metrics, config)
    item["metrics"] = metrics
    item["scores"] = scores
    item["recommendation"] = rec
    if include_points:
        item["points"] = sample_points(points, int(config.get("max_points_per_spectrum", 1200)))
    else:
        item["points"] = []
    return item


def mechanism_candidates(item: dict[str, Any] | None, availability: dict[str, Any], run_summary: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    metrics = (item or {}).get("metrics") or {}
    scores = (item or {}).get("scores") or {}
    flags = {flag.get("code") for flag in metrics.get("quality_flags", [])}
    has_r = availability.get("reflection", {}).get("present")
    has_a = availability.get("absorption", {}).get("present")
    has_field = availability.get("field", {}).get("present")
    has_phase = availability.get("phase", {}).get("present")
    has_poynting = availability.get("poynting", {}).get("present")
    candidates: list[dict[str, Any]] = []

    def add(name: str, confidence: float, support: list[str], missing: list[str], next_steps: list[str]):
        candidates.append(
            {
                "name": name,
                "claim": f"疑似 {name}",
                "confidence": round(clipped(confidence), 3),
                "supporting_evidence": support,
                "missing_evidence": missing,
                "next_steps": next_steps,
            }
        )

    if flags & {"severe_t_gt_1", "too_few_points", "fwhm_unreliable", "sampling_insufficient"}:
        support = []
        if "severe_t_gt_1" in flags:
            support.append("谱值明显超过 1.2，可能存在归一化、边界或收敛问题")
        if "too_few_points" in flags or "sampling_insufficient" in flags:
            support.append("采样点不足，窄峰/窄谷可能由少数点决定")
        if "fwhm_unreliable" in flags:
            support.append("FWHM 无法可靠计算")
        add(
            "数值伪影 / 采样不足",
            0.65 + 0.25 * clipped(float(scores.get("q_mode", 0)) / 100.0),
            support,
            ["缺少 mesh 收敛性、auto shutoff 和重复运行对比"],
            ["提高波长采样密度", "检查 auto shutoff、仿真时间和边界条件", "改变 mesh 后复跑候选点"],
        )

    if scores.get("fano", 0) > 35:
        add(
            "Fano 共振",
            float(scores.get("fano", 0)) / 100.0,
            ["峰谷相邻且非对称线形较明显", f"非对称因子约 {metrics.get('asymmetry', 0):.2f}"],
            [x for x, ok in (("相位跃迁证据", has_phase), ("场图模式分布", has_field), ("反射谱对应变化", has_r)) if not ok],
            ["导出共振附近相位谱", "在峰/谷两侧分别导出场图", "补充 R 谱并尝试 Fano 线型拟合"],
        )

    if scores.get("q_mode", 0) > 35 and (metrics.get("q") or 0) > 20:
        add(
            "准 BIC",
            min(0.78, float(scores.get("q_mode", 0)) / 100.0),
            [f"存在窄线宽高 Q 候选，Q≈{metrics.get('q'):.1f}" if metrics.get("q") else "存在窄线宽候选", "当前只能由 T 谱初筛"],
            ["缺少 Q-扰动强度标度关系", "缺少远场辐射或场图证据", "缺少高对称参考结构对比"],
            ["围绕候选参数加密扫描", "绘制 log(Q) 对 log(扰动强度) 的关系", "导出共振场图和远场方向图"],
        )

    if max(scores.get("notch", 0), scores.get("passband", 0), scores.get("q_mode", 0)) > 30:
        add(
            "导模共振 / GMR",
            clipped(max(scores.get("notch", 0), scores.get("passband", 0), scores.get("q_mode", 0)) / 120.0),
            ["出现相对窄带的透射峰/谷特征", "谱线对扫描参数可能具备可追踪中心波长"],
            [x for x, ok in (("反射谱增强证据", has_r), ("平面内延展场图", has_field), ("角度或周期敏感性扫描", False)) if not ok],
            ["补充 R 谱", "导出平面内场分布", "改变周期或入射角验证特征漂移"],
        )

    if metrics.get("num_peaks", 0) + metrics.get("num_dips", 0) >= 4:
        add(
            "Fabry-Perot 腔效应 / 多散射",
            0.38,
            ["谱线中存在多个峰/谷，可能来自多次干涉或集体耦合"],
            ["缺少厚度/间距扫描趋势", "缺少场图中驻波或结构间耦合证据"],
            ["扫描厚度、间距或周期参数", "导出多个共振波长的场图比较"],
        )

    if scores.get("broadband_low", 0) > 42 or scores.get("edge", 0) > 45:
        add(
            "Bragg / 光子带隙 或 瑞利异常",
            clipped(max(scores.get("broadband_low", 0), scores.get("edge", 0)) / 120.0),
            ["存在宽带低透或锐截止倾向"],
            ["缺少周期/衍射级次条件验证", "缺少角度扫描和反射谱"],
            ["检查特征波长与周期/折射率的关系", "补充角度扫描和 R 谱"],
        )

    if has_a:
        add(
            "吸收型共振",
            0.34,
            ["已发现吸收/损耗数据，可进一步检查 A 谱峰值"],
            [x for x, ok in (("材料损耗参数", False), ("场图中损耗区域能量集中", has_field)) if not ok],
            ["叠加 T/R/A 谱线", "导出有损材料区域的 |E|^2 分布"],
        )
    elif scores.get("notch", 0) > 38 and not has_r:
        add(
            "吸收型共振",
            0.22,
            ["T 谱存在低透射谷，但目前无法区分反射与吸收"],
            ["缺少 R 谱", "缺少 A 谱", "缺少有损材料区域场图"],
            ["补充 R 谱和 A 谱", "如果 A=1-T-R 较高，再导出损耗区域场图"],
        )

    note_text = " ".join(str((run_summary or {}).get(key, "")) for key in ("reduction_path", "relative_path", "perturbation"))
    if re.search(r"拓扑|topolog|bzf|fold", note_text, re.I):
        add(
            "拓扑相关模式变化",
            0.28,
            ["run 元数据或分析路径中包含拓扑/折叠相关线索"],
            ["缺少连续参数演化中的模式交换、反交叉或带隙闭合/重开证据"],
            ["在拓扑分析页查看热图演化", "围绕跃迁区域加密扫描并导出场图"],
        )

    if not candidates:
        add(
            "Mie 局域共振",
            0.18,
            ["当前仅凭 T 谱可作为低置信候选"],
            ["缺少结构内部局域场图", "缺少尺寸/折射率扫描趋势"],
            ["导出共振波长场图", "扫描尺寸或材料折射率验证中心波长漂移"],
        )
    candidates.sort(key=lambda item: item["confidence"], reverse=True)
    return candidates[:3]


def next_actions(best_item: dict[str, Any] | None, mechanisms: list[dict[str, Any]], missing: list[dict[str, Any]], scan_axis: dict[str, Any]) -> list[dict[str, Any]]:
    actions: list[dict[str, Any]] = []
    metrics = (best_item or {}).get("metrics") or {}
    if missing:
        for row in missing[:4]:
            actions.append({"priority": "P1" if row.get("key") in {"reflection", "absorption"} else "P2", "title": f"补充{row['label']}", "reason": row["why"], "detail": row["next"]})
    if metrics.get("fwhm_reliable") is False or metrics.get("sampling_insufficient"):
        actions.append(
            {
                "priority": "P1",
                "title": "加密候选波长附近采样",
                "reason": "当前 FWHM/Q 可能由过少采样点决定。",
                "detail": "围绕主峰/主谷中心波长设置更密集的 wavelength sweep，并保留原始 T/R/A 数据。",
            }
        )
    if best_item and best_item.get("scan_axis") and best_item.get("scan_value") is not None:
        actions.append(
            {
                "priority": "P2",
                "title": "围绕最佳扫描参数做局部复扫",
                "reason": "当前排行榜已出现相对高分候选。",
                "detail": f"以 {best_item.get('scan_axis')}≈{best_item.get('scan_value')} 为中心缩小步长，观察 λ0/FWHM/Q/score 是否连续演化。",
            }
        )
    for mech in mechanisms[:2]:
        steps = mech.get("next_steps") or []
        if steps:
            actions.append({"priority": "P2", "title": f"验证{mech['name']}", "reason": "机制判断目前仍是疑似结论，需要补证据。", "detail": "；".join(steps[:3])})
    if not actions:
        actions.append({"priority": "P3", "title": "保留当前结果并扩展目标函数", "reason": "当前数据不足以给出更强建议。", "detail": "先补齐 T/R/A 与关键场图，再比较不同目标类型的排行榜。"})
    return actions


def run_metadata(run_dir: Path, root: Path, run_summary: dict[str, Any] | None = None, register_file: Callable[[Path], str] | None = None) -> dict[str, Any]:
    rel = rel_path(run_dir, root)
    parts = Path(rel).parts
    if "results" in parts:
        idx = parts.index("results")
        group = " / ".join(parts[:idx])
        mother = parts[idx - 1] if idx >= 1 else ""
        perturbation = parts[idx + 1] if idx + 1 < len(parts) else ""
    else:
        group = " / ".join(parts[:-1])
        mother = parts[-2] if len(parts) >= 2 else ""
        perturbation = ""
    summary = run_summary or {}
    return {
        "id": summary.get("id", register_file(run_dir) if register_file else ""),
        "name": run_dir.name,
        "relative_path": rel,
        "group": summary.get("group", group),
        "mother": mother,
        "perturbation": summary.get("perturbation", perturbation),
        "reduction_path": summary.get("reduction_path", ""),
        "modified": run_dir.stat().st_mtime if run_dir.exists() else 0,
        "folder_id": register_file(run_dir) if register_file else summary.get("id", ""),
    }


def analysis_summary(items: list[dict[str, Any]], availability: dict[str, Any], target: str = "auto") -> dict[str, Any]:
    valid = [item for item in items if item.get("metrics", {}).get("point_count", 0) > 0]
    abnormal = [item for item in valid if any(flag.get("level") == "error" for flag in item.get("metrics", {}).get("quality_flags", []))]
    best = max(valid, key=lambda item: item.get("scores", {}).get("overall", 0.0), default=None)
    targets = [item.get("recommendation", {}).get("target") for item in valid]
    target_counts = {key: targets.count(key) for key in sorted(set(targets)) if key}
    return {
        "target": target,
        "spectrum_count": len(items),
        "valid_spectrum_count": len(valid),
        "abnormal_spectrum_count": len(abnormal),
        "high_value_count": sum(1 for item in valid if item.get("scores", {}).get("overall", 0.0) >= 60),
        "best_uid": best.get("uid") if best else "",
        "best_score": best.get("scores", {}).get("overall", 0.0) if best else 0.0,
        "best_target": best.get("recommendation", {}).get("target", "") if best else "",
        "best_lambda_nm": best.get("metrics", {}).get("center_lambda_nm") if best else None,
        "best_q": best.get("metrics", {}).get("q") if best else None,
        "target_counts": target_counts,
        "data_completeness": {key: {"present": value.get("present"), "count": value.get("count", 0)} for key, value in availability.items()},
    }


def rankings(items: list[dict[str, Any]], limit: int = 50) -> dict[str, list[dict[str, Any]]]:
    out: dict[str, list[dict[str, Any]]] = {}
    for target in ("overall", *TARGET_TYPES):
        rows = sorted(items, key=lambda item: item.get("scores", {}).get(target, 0.0), reverse=True)
        out[target] = [ranking_row(item, target) for item in rows[:limit]]
    return out


def ranking_row(item: dict[str, Any], target: str) -> dict[str, Any]:
    metrics = item.get("metrics") or {}
    scores = item.get("scores") or {}
    return {
        "uid": item.get("uid", ""),
        "name": item.get("name", ""),
        "file_name": item.get("file_name", ""),
        "scan_axis": item.get("scan_axis", ""),
        "scan_value": item.get("scan_value"),
        "target": item.get("recommendation", {}).get("target", ""),
        "score": scores.get(target, scores.get("overall", 0.0)),
        "overall": scores.get("overall", 0.0),
        "center_lambda_nm": metrics.get("center_lambda_nm"),
        "main_peak_lambda_nm": metrics.get("main_peak_lambda_nm"),
        "main_dip_lambda_nm": metrics.get("main_dip_lambda_nm"),
        "t_max": metrics.get("t_max"),
        "t_min": metrics.get("t_min"),
        "background_mean": metrics.get("background_mean"),
        "fwhm_nm": metrics.get("line_width_nm"),
        "q": metrics.get("q"),
        "side_peak_count": metrics.get("side_peak_count"),
        "side_dip_count": metrics.get("side_dip_count"),
        "quality_score": metrics.get("quality_score"),
        "flags": "; ".join(flag.get("code", "") for flag in metrics.get("quality_flags", [])),
        "file_id": item.get("file_id", ""),
        "png_id": item.get("png_id", ""),
    }


def strip_points(payload: dict[str, Any]) -> dict[str, Any]:
    data = json.loads(json.dumps(payload, ensure_ascii=False, default=str))
    for item in data.get("items", []):
        item.pop("points", None)
    return data


def csv_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, float):
        if not math.isfinite(value):
            return ""
        return f"{value:.8g}"
    return value


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow({field: csv_value(row.get(field, "")) for field in fields})


def format_number(value: Any, digits: int = 3) -> str:
    if value is None:
        return "不可可靠计算"
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)
    if not math.isfinite(number):
        return "不可可靠计算"
    return f"{number:.{digits}f}"


def diagnostic_markdown(payload: dict[str, Any]) -> str:
    run = payload.get("run", {})
    summary = payload.get("summary", {})
    best = None
    best_uid = summary.get("best_uid")
    for item in payload.get("items", []):
        if item.get("uid") == best_uid:
            best = item
            break
    lines = [
        f"# FDTD 光谱诊断摘要 - {run.get('name', '')}",
        "",
        f"- run: `{run.get('relative_path', '')}`",
        f"- 透射谱数量: {summary.get('spectrum_count', 0)}，有效谱线: {summary.get('valid_spectrum_count', 0)}，异常谱线: {summary.get('abnormal_spectrum_count', 0)}",
        f"- 最佳候选: {best_uid or '暂无'}，综合分: {format_number(summary.get('best_score'), 2)}，推荐目标: {summary.get('best_target') or '暂无'}",
    ]
    if best:
        metrics = best.get("metrics", {})
        lines.extend(
            [
                f"- 中心波长: {format_number(metrics.get('center_lambda_nm'), 3)} nm",
                f"- FWHM/线宽: {format_number(metrics.get('line_width_nm'), 3)} nm",
                f"- Q: {format_number(metrics.get('q'), 2)}",
                f"- T_max/T_min/背景: {format_number(metrics.get('t_max'), 4)} / {format_number(metrics.get('t_min'), 4)} / {format_number(metrics.get('background_mean'), 4)}",
            ]
        )
    lines.append("")
    lines.append("## 机制初判 Top 3")
    for mech in payload.get("mechanism_summary", {}).get("top", []):
        lines.append(f"- {mech.get('claim')}，置信度 {format_number(mech.get('confidence'), 2)}；支持证据：{'；'.join(mech.get('supporting_evidence', [])[:3])}")
    lines.append("")
    lines.append("## 缺失数据")
    for row in payload.get("missing_data", [])[:8]:
        lines.append(f"- {row.get('label')}: {row.get('why')} 下一步：{row.get('next')}")
    lines.append("")
    lines.append("## 下一轮建议")
    for action in payload.get("suggestions", []):
        lines.append(f"- [{action.get('priority')}] {action.get('title')}: {action.get('detail')}")
    lines.append("")
    return "\n".join(lines)


def write_analysis_outputs(run_dir: Path, payload: dict[str, Any], register_file: Callable[[Path], str] | None = None) -> dict[str, str]:
    out_dir = run_dir / "12_analysis_summary"
    out_dir.mkdir(parents=True, exist_ok=True)
    fields = [
        "uid",
        "name",
        "file_name",
        "scan_axis",
        "scan_value",
        "target",
        "overall",
        "center_lambda_nm",
        "main_peak_lambda_nm",
        "main_dip_lambda_nm",
        "t_max",
        "t_min",
        "background_mean",
        "fwhm_nm",
        "q",
        "side_peak_count",
        "side_dip_count",
        "quality_score",
        "flags",
    ]
    rows = [ranking_row(item, "overall") for item in payload.get("items", [])]
    metrics_csv = out_dir / "spectral_metrics.csv"
    write_csv(metrics_csv, rows, fields)
    metrics_json = out_dir / "spectral_metrics.json"
    metrics_json.write_text(json.dumps(strip_points(payload), ensure_ascii=False, indent=2), encoding="utf-8")
    for target in ("notch", "passband"):
        write_csv(out_dir / f"ranking_{target}.csv", payload.get("rankings", {}).get(target, []), list(payload.get("rankings", {}).get(target, [{}])[0].keys()) if payload.get("rankings", {}).get(target) else fields)
    mechanism_path = out_dir / "mechanism_guess.json"
    mechanism_path.write_text(json.dumps(payload.get("mechanism_summary", {}), ensure_ascii=False, indent=2), encoding="utf-8")
    missing_path = out_dir / "missing_data_report.json"
    missing_path.write_text(json.dumps(payload.get("missing_data", []), ensure_ascii=False, indent=2), encoding="utf-8")
    next_path = out_dir / "next_actions.md"
    next_lines = ["# 下一轮仿真建议", ""]
    for action in payload.get("suggestions", []):
        next_lines.append(f"- [{action.get('priority')}] {action.get('title')}")
        next_lines.append(f"  - 原因：{action.get('reason')}")
        next_lines.append(f"  - 建议：{action.get('detail')}")
    next_path.write_text("\n".join(next_lines) + "\n", encoding="utf-8")
    summary_path = out_dir / "diagnostic_summary.md"
    summary_path.write_text(diagnostic_markdown(payload), encoding="utf-8")
    files = {
        "spectral_metrics_csv": str(metrics_csv),
        "spectral_metrics_json": str(metrics_json),
        "ranking_notch_csv": str(out_dir / "ranking_notch.csv"),
        "ranking_passband_csv": str(out_dir / "ranking_passband.csv"),
        "mechanism_guess_json": str(mechanism_path),
        "missing_data_report_json": str(missing_path),
        "next_actions_md": str(next_path),
        "diagnostic_summary_md": str(summary_path),
    }
    if register_file:
        return {key: register_file(Path(value)) for key, value in files.items()}
    return files


def analyze_run_dir(
    run_dir: Path,
    root: Path,
    run_summary: dict[str, Any] | None = None,
    config: dict[str, Any] | None = None,
    register_file: Callable[[Path], str] | None = None,
    include_points: bool = True,
    write_outputs: bool = True,
) -> dict[str, Any]:
    config = merge_config(default_config(), config)
    if not run_dir.exists() or not run_dir.is_dir() or not is_under_root(run_dir, root):
        return {"ok": False, "error": "run directory not found or outside ROOT", "mode": "run"}

    availability = data_availability(run_dir, root, register_file)
    scan_map, manifest_map = load_run_rows(run_dir)
    trans_files, trans_images = discover_kind_files(run_dir, "transmission")
    items: list[dict[str, Any]] = []
    for path in trans_files:
        items.append(analyze_spectrum_item(path, "transmission", run_dir, root, scan_map, manifest_map, trans_images, config, register_file, include_points=include_points))
    items.sort(key=lambda item: natural_key(item.get("uid") or item.get("name") or ""))
    axis = infer_scan_axis(items)
    rank = rankings(items, int(config.get("global_top_n", 80)))
    summary = analysis_summary(items, availability, str(config.get("target_type", "auto")))
    best = max(items, key=lambda item: item.get("scores", {}).get("overall", 0.0), default=None)
    missing = missing_data_report(availability)
    mechanisms = mechanism_candidates(best, availability, run_summary)
    suggestions = next_actions(best, mechanisms, missing, axis)
    payload: dict[str, Any] = {
        "ok": True,
        "mode": "run",
        "generated_at": time.time(),
        "run": run_metadata(run_dir, root, run_summary, register_file),
        "summary": summary,
        "availability": availability,
        "scan_axis": axis,
        "items": items,
        "rankings": rank,
        "mechanism_summary": {"top": mechanisms, "basis": "保守启发式初判，仅作为下一步验证线索"},
        "missing_data": missing,
        "suggestions": suggestions,
        "config": config,
    }
    if write_outputs:
        try:
            payload["export_files"] = write_analysis_outputs(run_dir, payload, register_file)
        except Exception as exc:
            payload["export_error"] = str(exc)
    return payload


def cached_run_payload(run_dir: Path) -> dict[str, Any] | None:
    path = run_dir / "12_analysis_summary" / "spectral_metrics.json"
    if not path.exists():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None


def run_dir_from_summary(root: Path, run: dict[str, Any]) -> Path | None:
    rel = run.get("relative_path")
    if not rel:
        return None
    path = (root / str(rel)).resolve()
    if path.is_dir() and is_under_root(path, root):
        return path
    return None


def analyze_global(
    root: Path,
    scan_payload: dict[str, Any],
    config: dict[str, Any] | None = None,
    register_file: Callable[[Path], str] | None = None,
    force: bool = False,
) -> dict[str, Any]:
    config = merge_config(default_config(), config)
    run_payloads = []
    errors = []
    for run in scan_payload.get("runs", []):
        run_dir = run_dir_from_summary(root, run)
        if not run_dir:
            continue
        payload = None if force else cached_run_payload(run_dir)
        if not payload:
            payload = analyze_run_dir(run_dir, root, run, config, register_file, include_points=False, write_outputs=True)
        if payload.get("ok"):
            run_payloads.append(payload)
        elif payload.get("error"):
            errors.append({"run": run.get("relative_path", run.get("name", "")), "error": payload.get("error")})

    global_rows: list[dict[str, Any]] = []
    for payload in run_payloads:
        run = payload.get("run", {})
        for row in payload.get("rankings", {}).get("overall", [])[: max(1, min(5, int(config.get("global_top_n", 80))))]:
            merged = dict(row)
            merged.update(
                {
                    "run_id": run.get("id", ""),
                    "run_name": run.get("name", ""),
                    "run_path": run.get("relative_path", ""),
                    "group": run.get("group", ""),
                    "perturbation": run.get("perturbation", ""),
                    "folder_id": run.get("folder_id", ""),
                }
            )
            global_rows.append(merged)

    def top_for(target: str) -> list[dict[str, Any]]:
        rows = []
        for payload in run_payloads:
            run = payload.get("run", {})
            for row in payload.get("rankings", {}).get(target, [])[:5]:
                merged = dict(row)
                merged.update({"run_id": run.get("id", ""), "run_name": run.get("name", ""), "run_path": run.get("relative_path", ""), "group": run.get("group", ""), "perturbation": run.get("perturbation", ""), "folder_id": run.get("folder_id", "")})
                rows.append(merged)
        return sorted(rows, key=lambda item: float(item.get("score") or item.get("overall") or 0.0), reverse=True)[: int(config.get("global_top_n", 80))]

    rankings_global = {target: top_for(target) for target in ("overall", "notch", "passband", "fano", "q_mode", "edge", "broadband_high", "broadband_low", "flat")}
    summaries = [payload.get("summary", {}) for payload in run_payloads]
    summary = {
        "run_count": len(run_payloads),
        "spectrum_count": sum(int(s.get("spectrum_count") or 0) for s in summaries),
        "valid_spectrum_count": sum(int(s.get("valid_spectrum_count") or 0) for s in summaries),
        "abnormal_spectrum_count": sum(int(s.get("abnormal_spectrum_count") or 0) for s in summaries),
        "high_value_count": sum(int(s.get("high_value_count") or 0) for s in summaries),
        "missing_reflection_runs": sum(1 for payload in run_payloads if not payload.get("availability", {}).get("reflection", {}).get("present")),
        "missing_field_runs": sum(1 for payload in run_payloads if not payload.get("availability", {}).get("field", {}).get("present")),
        "errors": errors[:20],
    }
    return {
        "ok": True,
        "mode": "global",
        "generated_at": time.time(),
        "root": str(root),
        "summary": summary,
        "rankings": rankings_global,
        "runs": [
            {
                "run": payload.get("run", {}),
                "summary": payload.get("summary", {}),
                "availability": payload.get("availability", {}),
                "missing_data": payload.get("missing_data", []),
            }
            for payload in run_payloads
        ],
        "config": config,
    }


def write_global_exports(root: Path, payload: dict[str, Any], register_file: Callable[[Path], str] | None = None) -> dict[str, str]:
    out_dir = root / "12_spectral_global_summary"
    out_dir.mkdir(parents=True, exist_ok=True)
    fields = [
        "run_name",
        "run_path",
        "group",
        "perturbation",
        "uid",
        "name",
        "target",
        "score",
        "overall",
        "center_lambda_nm",
        "t_max",
        "t_min",
        "fwhm_nm",
        "q",
        "quality_score",
        "flags",
    ]
    top_path = out_dir / "global_top_candidates.csv"
    write_csv(top_path, payload.get("rankings", {}).get("overall", []), fields)
    json_path = out_dir / "global_spectral_diagnostics.json"
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    files = {"global_top_csv": str(top_path), "global_json": str(json_path)}
    if register_file:
        return {key: register_file(Path(value)) for key, value in files.items()}
    return files
