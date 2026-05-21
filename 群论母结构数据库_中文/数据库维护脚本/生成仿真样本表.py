# -*- coding: utf-8 -*-
"""根据《对称扰动规则表.xlsx》生成《仿真样本参数表.xlsx》的范例批量样本。"""
from pathlib import Path
import json
import openpyxl

数据库根目录 = Path(__file__).resolve().parents[1]
表格目录 = 数据库根目录 / "数据表格模板"
规则表路径 = 表格目录 / "对称扰动规则表.xlsx"
样本表路径 = 表格目录 / "仿真样本参数表.xlsx"

样本表头 = ['仿真样本编号','母结构家族编号','扰动规则编号','结构中文名','扰动前对称群','扰动后对称群','降群路径','归一化扰动delta','绝对扰动_nm','几何参数_JSON','周期_nm','厚度_nm','谐振体材料','衬底材料','背景材料','晶格类型','入射偏振','波长下限_nm','波长上限_nm','x边界条件','y边界条件','z边界条件','网格尺寸_nm','透射监视器名称','仿真软件','仿真状态','创建日期','备注']
结构名映射 = {'C4_cross': '十字', 'C2_dual_pillars': '双柱', 'C2_dual_disks': '双圆盘', 'C4_square_ring': '方环'}

def 读取第一张表(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    return [dict(zip(headers, row)) for row in rows[1:] if any(row)]

def 写入工作簿(path, headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '仿真样本参数表'
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, '') for h in headers])
    ws.freeze_panes = 'A2'
    wb.save(path)

def 生成几何参数(家族编号, delta):
    周期, 厚度 = 500, 180
    if 家族编号 == 'C4_cross':
        L0 = 240
        Lx = L0 * (1 + delta / 2)
        Ly = L0 * (1 - delta / 2)
        return {'period_p_nm':周期,'h_nm':厚度,'L0_nm':L0,'Lx_nm':round(Lx,3),'Ly_nm':round(Ly,3),'Wx_nm':60,'Wy_nm':60}, abs(Lx-Ly)
    return {'period_p_nm':周期,'h_nm':厚度}, 0

规则列表 = 读取第一张表(规则表路径)
样本列表 = []
for 规则 in 规则列表:
    家族编号 = 规则.get('母结构家族编号')
    if not 家族编号:
        continue
    delta文本 = str(规则.get('推荐delta序列') or '0')
    for 序号, delta in enumerate([float(x) for x in delta文本.split(',') if x.strip()], 1):
        几何参数, 绝对扰动 = 生成几何参数(家族编号, delta)
        样本编号 = f"{规则.get('扰动规则编号')}_{序号:04d}"
        样本列表.append({'仿真样本编号':样本编号,'母结构家族编号':家族编号,'扰动规则编号':规则.get('扰动规则编号'),'结构中文名':结构名映射.get(家族编号,''),'扰动前对称群':规则.get('扰动前对称群'),'扰动后对称群':规则.get('扰动后对称群'),'降群路径':规则.get('降群路径'),'归一化扰动delta':delta,'绝对扰动_nm':round(绝对扰动,3),'几何参数_JSON':json.dumps(几何参数,ensure_ascii=False,separators=(',',':')),'周期_nm':500,'厚度_nm':180,'谐振体材料':'TiO2','衬底材料':'SiO2','背景材料':'air','晶格类型':'正方晶格','入射偏振':'Ex','波长下限_nm':400,'波长上限_nm':900,'x边界条件':'周期边界','y边界条件':'周期边界','z边界条件':'PML','网格尺寸_nm':2,'透射监视器名称':'Trans','仿真软件':'Lumerical FDTD','仿真状态':'planned','创建日期':'','备注':'脚本自动生成；每次只改变一种扰动强度'})
写入工作簿(样本表路径, 样本表头, 样本列表)
print(f"已生成 {len(样本列表)} 个样本：{样本表路径}")
